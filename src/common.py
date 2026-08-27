# -*- coding: utf-8 -*-
import json, math, re, statistics as st
from collections import Counter, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule

D = json.load(open("core.json", encoding="utf-8"))
S = json.load(open("stats.json", encoding="utf-8"))
Q, P, CORR = D["Q"], D["P"], D["CORR"]
TIME = {n: {int(k): v for k, v in d.items()} for n, d in D["TIME"].items()}
items = S["items"]
N = 37

NAVY = "1F3864"; BLUE = "2E5C9A"; LGREY = "EEF2F8"


def F(c):
    return PatternFill("solid", fgColor=c)


HDR = F(NAVY); SUBF = F(BLUE)
WF = Font(color="FFFFFF", bold=True, size=10)
TH = Side(style="thin", color="C5CEDC")
BOX = Border(TH, TH, TH, TH)
GRN = F("C6EFCE"); RED = F("FFC7CE"); YEL = F("FFEB9C"); GRY = F("DDDDDD")


def head(ws, row, cols, widths=None):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row, i, c)
        cell.fill = HDR; cell.font = WF; cell.border = BOX
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[gcl(i)].width = w
    ws.row_dimensions[row].height = 34


def title(ws, text, sub="", span=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, text)
    c.fill = HDR; c.font = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c = ws.cell(2, 1, sub)
    c.fill = F(LGREY); c.font = Font(size=9, italic=True, color="44546A")
    c.alignment = Alignment(horizontal="left", indent=1, vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 34


def band(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = BOX


def secrow(ws, r, text, span):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    c = ws.cell(r, 1, text)
    c.fill = SUBF; c.font = Font(color="FFFFFF", bold=True, size=10)
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[r].height = 20


def note(ws, r, text, span=8, h=40, bold=False, size=10):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    c = ws.cell(r, 1, text)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(size=size, bold=bold)
    c.border = BOX
    ws.row_dimensions[r].height = h


sc = {n: P[n]["correct"] for n in P}
sco = {n: P[n]["score"] for n in P}
tt = {n: P[n]["time_s"] for n in P}
vals = sorted(sc.values())
mean_s = st.mean(sc.values())
sd_s = st.pstdev(sc.values())


def q_(p):
    i = (len(vals) - 1) * p
    lo = int(i); hi = min(lo + 1, len(vals) - 1)
    return vals[lo] + (vals[hi] - vals[lo]) * (i - lo)


Q1v, Q2v, Q3v = q_(.25), q_(.5), q_(.75)
RANK = sorted(P, key=lambda n: (-sco[n], -sc[n]))


def base(n):
    return re.sub(r"\*+$", "", n).strip()


def grade(pctg):
    for lim, g, l in [(85, "A", "Sangat Baik"), (70, "B", "Baik"), (55, "C", "Cukup"),
                      (40, "D", "Kurang"), (0, "E", "Sangat Kurang")]:
        if pctg >= lim:
            return g, l


def zof(n):
    return (sc[n] - mean_s) / sd_s


def prank(n):
    below = sum(1 for m in P if sc[m] < sc[n])
    same = sum(1 for m in P if sc[m] == sc[n])
    return (below + 0.5 * same) / N * 100


# domain / bloom mapping
DOMAIN = {
    1: "Akses & Model Bisnis", 3: "Akses & Model Bisnis",
    2: "Konsep & Fitur Canva", 4: "Konsep & Fitur Canva", 5: "Konsep & Fitur Canva",
    9: "Konsep & Fitur Canva", 12: "Konsep & Fitur Canva", 16: "Konsep & Fitur Canva",
    18: "Konsep & Fitur Canva",
    6: "Prinsip Desain Grafis", 8: "Prinsip Desain Grafis", 11: "Prinsip Desain Grafis",
    15: "Prinsip Desain Grafis",
    7: "Teknis Output & Ukuran", 14: "Teknis Output & Ukuran",
    10: "Sejarah & Profil", 20: "Sejarah & Profil",
    13: "Materi Internal Pelatihan", 17: "Materi Internal Pelatihan", 19: "Materi Internal Pelatihan",
}
BLOOM = {
    1: "C3 Menerapkan", 3: "C3 Menerapkan", 7: "C3 Menerapkan",
    2: "C2 Memahami", 4: "C2 Memahami", 5: "C2 Memahami", 6: "C2 Memahami", 8: "C2 Memahami",
    9: "C2 Memahami", 11: "C2 Memahami", 12: "C2 Memahami", 15: "C2 Memahami", 16: "C2 Memahami",
    18: "C2 Memahami",
    10: "C1 Mengingat", 13: "C1 Mengingat", 14: "C1 Mengingat", 17: "C1 Mengingat",
    19: "C1 Mengingat", 20: "C1 Mengingat",
}
SHORT = {
    1: "Langkah pertama memakai Canva gratis",
    2: "Pernyataan benar tentang Canva",
    3: "Rekomendasi paket harga",
    4: "Cara kerja drag-and-drop",
    5: "Fungsi Panel Kiri vs Area Desain",
    6: "Pelanggaran hierarki visual",
    7: "Format ekspor untuk Instagram",
    8: "Kombinasi warna analog",
    9: "Beda Template vs Elemen",
    10: "Bisnis asal-usul Canva",
    11: "Makna White Space",
    12: "Manfaat penyimpanan cloud",
    13: "Yang BUKAN tujuan pelatihan",
    14: "Ukuran TikTok / IG Story",
    15: "Praktik penggunaan font",
    16: "Manfaat kolaborasi tim",
    17: "Slogan pelatihan",
    18: "Urutan menambah elemen",
    19: "Yang BUKAN nilai Berkarya dengan Hati",
    20: "Jabatan pemateri",
}


def dcat(d):
    if d >= .40: return "Sangat Baik", "C6EFCE"
    if d >= .30: return "Baik", "D9EAD3"
    if d >= .20: return "Cukup / revisi", "FFEB9C"
    if d > 0: return "Buruk / revisi besar", "FCE4D6"
    return "Ditolak", "FFC7CE"


def pcat(p):
    if p >= .70: return "Mudah", "C6EFCE"
    if p >= .30: return "Sedang", "FFEB9C"
    return "Sukar", "FFC7CE"
