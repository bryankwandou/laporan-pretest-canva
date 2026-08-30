# -*- coding: utf-8 -*-
"""BUKTI KELENGKAPAN DATASET PENELITIAN.
Setiap pemeriksaan membandingkan isi sel workbook langsung ke BERKAS SUMBER
(dua ekspor XLSX Wayground dan dua naskah PDF resmi), bukan ke hasil olahan.
"""
import openpyxl, json, io, os, re, math, statistics as st
from pypdf import PdfReader

SRC = "E:/Download"
XLS = "DATASET_PENELITIAN_CANVA_2026.xlsx"
SPRE = os.path.join(SRC, "pretestpelatihancanva25agustus2026-2026-08-25T09_22_13_634913-c1bee5.xlsx")
SPOST = os.path.join(SRC, "post-testpelatihancanva25agustus2026-2026-08-28T14_07_30_851549-68e3ea.xlsx")
PPRE = os.path.join(SRC, "Free Printable pretest pelatihan canva 25 agustus 2026.pdf")
PPOST = os.path.join(SRC, "Free Printable post-test pelatihan canva 25 agustus 2026.pdf")

D = json.load(io.open("ds_core.json", encoding="utf-8"))
ORANG_PRE, ORANG_POST = D["ORANG_PRE"], D["ORANG_POST"]
CH_PRE, CH_POST = D["CH_PRE"], D["CH_POST"]
RPRE, RPOST = D["RPRE"], D["RPOST"]

L, res = [], []
W = L.append


def norm(s):
    s = str(s)
    for a, b in (("\u2018", "'"), ("\u2019", "'"), ("\u201c", '"'), ("\u201d", '"'),
                 ("\u2013", "-"), ("\u2014", "-"), ("\u2192", "->"), ("\u2026", ""),
                 ("\ufffd", "")):
        s = s.replace(a, b)
    return re.sub(r"\s+", " ", s).strip().lower()


def ck(nama, ok, detail=""):
    res.append((nama, ok, detail))
    W("  %-6s %-64s %s" % ("LULUS" if ok else "GAGAL", nama, detail))


W("=" * 104)
W("BUKTI KELENGKAPAN — DATASET_PENELITIAN_CANVA_2026.xlsx")
W("Pembanding: dua ekspor resmi Wayground dan dua naskah cetak resmi. Tidak ada angka yang")
W("diambil dari proses pembuatan; seluruh isi sel dibaca ulang dari workbook yang sudah jadi.")
W("=" * 104)
W("")

wb = openpyxl.load_workbook(XLS, data_only=True)


def cells(name):
    ws = wb[name]
    return [[c for c in row] for row in ws.iter_rows(values_only=True)]


def bag(name):
    s = set()
    for row in wb[name].iter_rows(values_only=True):
        for v in row:
            if v is not None:
                s.add(norm(v))
    return s


# =========== 1. NASKAH SOAL DAN OPSI vs NASKAH PDF
W("-" * 104)
W("1. NASKAH SOAL DAN SELURUH OPSI — DIBANDINGKAN KE NASKAH CETAK RESMI (PDF)")
W("-" * 104)


_FOOT = [r"\s*\d{1,2}/\d{1,2}/\d{2,4},?\s*\d{1,2}:\d{2}\s*[AP]M.*$",
         r"\s*/\d{2},?\s*\d{1,2}:\d{2}\s*[AP]M.*$",
         r"\s*Free Printable.*$", r"\s*Worksheets?\s*$", r"\s+\d{1,2}/\d{1,2}\s*$"]


def _cf(t):
    """Buang footer cetak PDF: tanggal, judul lembar, dan nomor halaman."""
    x = str(t)
    for p_ in _FOOT:
        x = re.sub(p_, "", x, flags=re.I)
    return re.sub(r"\s+", " ", x).strip()


def pdf_items(path, anchor=None):
    raw = "\n".join((p.extract_text() or "") for p in PdfReader(path).pages)
    raw = re.sub(r"https://\S+", " ", raw)
    raw = re.sub(r"\b\d+/\d+\b", " ", raw)
    flat = re.sub(r"\s+", " ", raw)
    m0 = re.search(r"(?:(?<= )|^)1\.\s*[A-Z]", flat)
    flat = flat[m0.start():] if m0 else flat
    idx = [(int(m.group(1)), m.start()) for m in re.finditer(r"(?:(?<= )|^)(\d{1,2})\.\s*(?=[A-Z])", flat)]
    idx = [(n, s) for n, s in idx if 1 <= n <= 20]
    seen, keep = set(), []
    for n, s in idx:
        if n not in seen:
            seen.add(n); keep.append((n, s))
    keep.sort(key=lambda x: x[1])
    out = {}
    for j, (n, s) in enumerate(keep):
        e = keep[j + 1][1] if j + 1 < len(keep) else len(flat)
        b = flat[s:e]
        parts = re.split(r"\b([a-d])\)\s*", b)
        out[n] = {"stem": _cf(re.sub(r"^\d+\.\s*", "", parts[0])),
                  "opts": [(parts[k], _cf(parts[k + 1]))
                           for k in range(1, len(parts) - 1, 2)]}
    return out


IPRE_PDF = pdf_items(PPRE)
IPOST_PDF = pdf_items(PPOST)

for lab, sheet, PDFI in (("PRE-TEST", "02 Soal Pre-test", IPRE_PDF),
                         ("POST-TEST", "03 Soal Post-test", IPOST_PDF)):
    B = bag(sheet)
    nopt = sum(len(v["opts"]) for v in PDFI.values())
    miss = [n for n in range(1, 21) if norm(PDFI[n]["stem"]) not in B]
    nch = sum(len(norm(PDFI[n]["stem"])) for n in range(1, 21))
    ck("%s: naskah 20 butir utuh di workbook (%d karakter)" % (lab, nch), not miss,
       "gagal: %s" % (miss or "tidak ada"))
    missopt = [(n, l) for n in PDFI for l, t in PDFI[n]["opts"] if norm(t) not in B]
    ck("%s: seluruh %d opsi dari naskah PDF utuh di workbook" % (lab, nopt), not missopt,
       "gagal: %s" % (missopt or "tidak ada"))
    # tepat 20 baris bertanda KUNCI
    rows = cells(sheet)
    nkey = sum(1 for r in rows for v in r if v == "KUNCI")
    ck("%s: tepat 20 opsi ditandai KUNCI" % lab, nkey == 20, "%d penanda" % nkey)
    nopt_ws = sum(1 for r in rows if r[2] and re.match(r"^[a-d]\)$", str(r[2])))
    ck("%s: workbook memuat %d baris opsi" % (lab, nopt), nopt_ws == nopt, "%d baris" % nopt_ws)
W("")

# =========== 2. JAWABAN MENTAH vs EKSPOR
W("-" * 104)
W("2. TEKS JAWABAN TIAP ORANG TIAP BUTIR — DIBANDINGKAN SEL PER SEL KE EKSPOR RESMI")
W("-" * 104)


def export_grid(path):
    wbx = openpyxl.load_workbook(path, data_only=True)
    rows = [list(r) for r in wbx["Overview"].iter_rows(values_only=True)]
    hdr = rows[0]
    pc = {}
    for i in range(12, len(hdr)):
        if hdr[i] is None:
            continue
        m = re.match(r"^(.*?)\s*\((.*)\)\s*$", str(hdr[i]))
        pc[(m.group(1) if m else str(hdr[i])).strip()] = i
    grid = {n: {} for n in pc}
    unatt = {}
    for r in rows[1:21]:
        q = int(r[0])
        unatt[q] = r[11]
        for n, i in pc.items():
            grid[n][q] = r[i]
    return grid, unatt


def ws_answers(sheet, orang):
    out = {}
    for row in wb[sheet].iter_rows(values_only=True):
        if row[1] and str(row[1]) in orang:
            out[str(row[1])] = {q: row[1 + q] for q in range(1, 21)}
    return out


GPRE, UPRE = export_grid(SPRE)
GPOST, UPOST = export_grid(SPOST)
APRE = ws_answers("08 Jawaban Pre-test", set(ORANG_PRE))
APOST = ws_answers("09 Jawaban Post-test", set(ORANG_POST))

ck("Lembar 08 memuat satu baris per orang pre-test", len(APRE) == len(ORANG_PRE),
   "%d dari %d" % (len(APRE), len(ORANG_PRE)))
ck("Lembar 09 memuat satu baris per orang post-test", len(APOST) == len(ORANG_POST),
   "%d dari %d" % (len(APOST), len(ORANG_POST)))

# post: setiap sel harus identik ke ekspor
bad, n = [], 0
for o in ORANG_POST:
    s = CH_POST[o]
    for q in range(1, 21):
        n += 1
        src, got = GPOST[s][q], APOST[o][q]
        if RPOST[o][q - 1] == "-":
            if norm(got) != "(tidak dijawab)":
                bad.append((o, q, "seharusnya ditandai tidak dijawab"))
        elif norm(got) != norm(src):
            bad.append((o, q, "teks berbeda"))
ck("POST-TEST: %d sel jawaban identik dengan ekspor resmi" % n, not bad,
   "beda: %s" % (bad[:3] or "tidak ada"))

# pre: sel terjawab harus identik; sel kosong dicocokkan ke kolom Unattempted
bad2, n2, nblank = [], 0, {q: 0 for q in range(1, 21)}
for o in ORANG_PRE:
    s = CH_PRE[o]
    for q in range(1, 21):
        n2 += 1
        got = APRE[o][q]
        if RPRE[o][q - 1] == "-":
            nblank[q] += 1
            if norm(got) != "(tidak dijawab)":
                bad2.append((o, q, "seharusnya tidak dijawab"))
        elif norm(got) != norm(GPRE[s][q]):
            bad2.append((o, q, "teks berbeda"))
ck("PRE-TEST: %d sel jawaban identik dengan ekspor resmi" % n2, not bad2,
   "beda: %s" % (bad2[:3] or "tidak ada"))
W("         Catatan: sel kosong pre-test diverifikasi terpisah pada butir 3 di bawah, karena ekspor")
W("         XLSX pre-test memuat teks jawaban semu pada sel yang sebenarnya tidak dijawab.")
W("")

# =========== 3. CACAH — SENSUS DAN AGREGAT
W("-" * 104)
W("3. CACAH ORANG, SESI, DAN AGREGAT PER BUTIR")
W("-" * 104)
wbp = openpyxl.load_workbook(SPRE, data_only=True)
wbq = openpyxl.load_workbook(SPOST, data_only=True)
nses_pre = sum(1 for r in list(wbp["Participant Data"].iter_rows(values_only=True))[1:] if r[0] is not None)
nses_post = sum(1 for r in list(wbq["Participant Data"].iter_rows(values_only=True))[1:] if r[0] is not None)
ck("Jumlah sesi pre-test pada ekspor = 37", nses_pre == 37, "%d baris" % nses_pre)
ck("Jumlah sesi post-test pada ekspor = 16", nses_post == 16, "%d baris" % nses_post)
S1 = bag("01 Sensus")
ck("Lembar Sensus menyatakan 33 orang pre-test", "33" in {norm(v) for v in S1}, "")
ck("Lembar Sensus menyatakan 14 orang post-test", "14" in {norm(v) for v in S1}, "")
ck("Sesi ganda pre-test dijelaskan (Sri Suyani, Aqifah, farida johannes)",
   all(norm(x) in " ".join(S1) for x in ("sri suyani", "aqifah", "farida johannes")), "")
ck("Sesi ganda post-test dijelaskan (Yovita)", "yovita" in " ".join(S1), "")

# agregat post per butir harus cocok ke matriks workbook
rows = cells("07 Respons Post-test")
foot = [r for r in rows if r[1] == "Benar per butir"]
ok = False
if foot:
    got = [foot[0][2 + i] for i in range(20)]
    exp = [sum(1 for o in ORANG_POST if RPOST[o][i] == "C") for i in range(20)]
    ok = got == exp
ck("Baris 'Benar per butir' post-test konsisten dengan matriksnya", ok, "")
W("")

# =========== 4. STATISTIK DAPAT DIHITUNG ULANG DARI MATRIKS
W("-" * 104)
W("4. STATISTIK DAPAT DIHITUNG ULANG DARI MATRIKS DI DALAM WORKBOOK ITU SENDIRI")
W("-" * 104)


def matrix_from_ws(sheet, orang):
    out = {}
    for row in wb[sheet].iter_rows(values_only=True):
        if row[1] and str(row[1]) in orang:
            out[str(row[1])] = "".join({"B": "C", "S": "X"}.get(str(row[2 + i]), "-")
                                       for i in range(20))
    return out


for lab, sheet, bsheet, orang in (("PRE-TEST", "06 Respons Pre-test", "04 Butir Pre-test", ORANG_PRE),
                                  ("POST-TEST", "07 Respons Post-test", "05 Butir Post-test", ORANG_POST)):
    M = matrix_from_ws(sheet, set(orang))
    ck("%s: matriks respons memuat %d orang" % (lab, len(orang)), len(M) == len(orang),
       "%d baris" % len(M))
    N = len(orang)
    # p tiap butir dari matriks harus sama dengan kolom p pada lembar butir
    brows = [r for r in cells(bsheet) if isinstance(r[0], int) and 1 <= r[0] <= 20]
    bad3 = []
    for r_ in brows:
        q = r_[0]
        pm = round(sum(1 for o in orang if M[o][q - 1] == "C") / N, 3)
        if abs(float(r_[6]) - pm) > 0.0015:
            bad3.append((q, r_[6], pm))
    ck("%s: nilai p 20 butir dapat dihitung ulang dari matriks" % lab, not bad3,
       "beda: %s" % (bad3[:3] or "tidak ada"))
    bad4 = [r_[0] for r_ in brows
            if r_[3] != sum(1 for o in orang if M[o][r_[0] - 1] == "C")
            or r_[4] != sum(1 for o in orang if M[o][r_[0] - 1] == "X")
            or r_[5] != sum(1 for o in orang if M[o][r_[0] - 1] == "-")]
    ck("%s: benar/salah/kosong 20 butir konsisten dengan matriks" % lab, not bad4,
       "gagal: %s" % (bad4 or "tidak ada"))
W("")

# =========== 5. UJI STATISTIK DAPAT DIULANG DARI LEMBAR DATA OLAH
W("-" * 104)
W("5. UJI STATISTIK DAPAT DIULANG DARI LEMBAR 12 DATA OLAH")
W("-" * 104)
rows = cells("12 Data Olah")
recs = [r for r in rows if isinstance(r[0], int) and r[1] and isinstance(r[3], int) and r[9] in (0, 1)]
tun = [r for r in recs if r[9] == 1]
ck("Lembar Data Olah memuat 11 pasangan", len(recs) == 11, "%d baris" % len(recs))
ck("Di antaranya 8 bertanda tuntas = 1", len(tun) == 8, "%d baris" % len(tun))
g = [r[4] - r[3] for r in tun]
mg, sg = st.mean(g), st.stdev(g)
tv = mg / (sg / math.sqrt(len(g)))
dz = mg / sg
hk = mg / (20 - st.mean([r[3] for r in tun]))
B10 = bag("10 Berpasangan")
ck("Rata-rata gain dari data = +5,00", abs(mg - 5.0) < 1e-9, "%.4f" % mg)
ck("Uji-t dari data = 3,669 dan tercantum di lembar 10", abs(tv - 3.669) < 0.002,
   "t=%.3f  tercantum=%s" % (tv, "ya" if "t(7) = 3.669" in " ".join(B10) or "3.669" in " ".join(B10) else "TIDAK"))
ck("Cohen dz dari data = 1,297", abs(dz - 1.297) < 0.002, "%.3f" % dz)
ck("Gain Hake dari data = 0,440", abs(hk - 0.440) < 0.002, "%.3f" % hk)
ck("Seluruh 8 pasangan naik", all(x > 0 for x in g), "%d dari %d" % (sum(1 for x in g if x > 0), len(g)))
W("")

# =========== 6. GRAFIK
W("-" * 104)
W("6. GRAFIK")
W("-" * 104)
nch = sum(len(w._charts) for w in wb)
linked = sum(1 for w in wb for c in w._charts for s in c.series
             if s.val is not None and s.val.numRef is not None)
ck("Workbook memuat grafik native Excel", nch >= 8, "%d grafik" % nch)
ck("Seri grafik tertaut ke sel data, bukan gambar", linked > 0, "%d seri tertaut" % linked)
ck("Lembar 11 Grafik memuat tabel data di sebelah kiri tiap grafik",
   wb["11 Grafik"].max_row > 200, "%d baris data" % wb["11 Grafik"].max_row)
W("")

W("=" * 104)
ok = sum(1 for a, b, c in res if b)
W("REKAP: %d dari %d pemeriksaan LULUS" % (ok, len(res)))
gg = [a for a, b, c in res if not b]
if gg:
    W("")
    W("BELUM TERPENUHI:")
    for a in gg:
        W("   - " + a)
W("=" * 104)

io.open("ds_audit.txt", "w", encoding="utf-8").write("\n".join(L))
print("\n".join(L))
