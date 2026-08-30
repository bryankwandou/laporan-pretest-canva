# -*- coding: utf-8 -*-
"""Workbook penelitian bersih — berbasis ORANG. Bagian A: kerangka + sheet 00-01."""
import json, io, math, re, statistics as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList

D = json.load(io.open("ds_core.json", encoding="utf-8"))
pre = json.load(io.open("core.json", encoding="utf-8"))
post = json.load(io.open("post_core.json", encoding="utf-8"))
Spre = json.load(io.open("stats.json", encoding="utf-8"))
OPRE = json.load(io.open("pdf_opts_pre.json", encoding="utf-8"))
OPOST = json.load(io.open("pdf_opts.json", encoding="utf-8"))
TJ = json.load(io.open("post_time.json", encoding="utf-8"))

ORANG_PRE, ORANG_POST = D["ORANG_PRE"], D["ORANG_POST"]
RPRE, RPOST = D["RPRE"], D["RPOST"]
CH_PRE, CH_POST = D["CH_PRE"], D["CH_POST"]
SPRE_I = {int(k): v for k, v in D["SPRE_I"].items()}
SPOST_I = {int(k): v for k, v in D["SPOST_I"].items()}
PAIR, PAIRT, STAT = D["PAIR"], D["PAIRT"], D["stat"]
NPRE, NPOST = len(ORANG_PRE), len(ORANG_POST)

QPRE = {q["no"]: q for q in pre["Q"]}
QPOST = {q["no"]: q for q in post["Q"]}
IPRE = {i["no"]: i for i in Spre["items"]}
OPT_PRE = {int(k): v for k, v in OPRE["OPT"].items()}
OPT_POST = {int(k): v for k, v in OPOST["OPT"].items()}
LET_PRE = {int(k): v for k, v in OPRE["LET"].items()}
DEAD_PRE = {int(k): v for k, v in OPRE["DEAD"].items()}
DEAD_POST = {int(k): v for k, v in OPOST["DEAD"].items()}
TIME = {n: {int(k): v for k, v in d.items()} for n, d in TJ["TIME"].items()}
PP, PO = pre["P"], post["P"]

NAVY = "12284B"; BLUE = "2F6FB5"; LGREY = "EDF0F5"
GRN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YEL = PatternFill("solid", fgColor="FFEB9C")
GRY = PatternFill("solid", fgColor="E8EAED")
BLU = PatternFill("solid", fgColor="DDEBF7")
HDR = PatternFill("solid", fgColor=NAVY)
SUB = PatternFill("solid", fgColor=BLUE)
LG = PatternFill("solid", fgColor=LGREY)
WF = Font(color="FFFFFF", bold=True, size=10)
TH = Side(style="thin", color="BFBFBF")
BOX = Border(TH, TH, TH, TH)
CEN = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(vertical="top", wrap_text=True)

wb = openpyxl.Workbook()
wb.remove(wb.active)


def title(ws, t, sub, span):
    c = ws.cell(1, 1, t)
    c.font = Font(bold=True, size=14, color=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(2, 1, sub)
    c.font = Font(size=9, italic=True, color="4A5568")
    c.alignment = WRAP
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    ws.row_dimensions[2].height = 28


def head(ws, r, cols, widths=None, start=1):
    for i, t in enumerate(cols):
        c = ws.cell(r, start + i, t)
        c.fill = HDR; c.font = WF; c.border = BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths):
            ws.column_dimensions[gcl(start + i)].width = w
    ws.row_dimensions[r].height = 30
    return r + 1


def sec(ws, r, t, span):
    c = ws.cell(r, 1, t)
    c.fill = SUB; c.font = Font(bold=True, size=10, color="FFFFFF")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    return r + 1


def note(ws, r, t, span, h=26):
    c = ws.cell(r, 1, t)
    c.font = Font(size=9, italic=True)
    c.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    ws.row_dimensions[r].height = h
    return r + 1


def row(ws, r, vals, ctr=(), bold=(), h=18, fs=10, start=1):
    for i, v in enumerate(vals):
        c = ws.cell(r, start + i, v)
        c.border = BOX
        c.font = Font(size=fs, bold=(i + 1) in bold)
        c.alignment = CEN if (i + 1) in ctr else WRAP
    ws.row_dimensions[r].height = h
    return r + 1


# ================= 00 PANDUAN
ws = wb.create_sheet("00 Panduan")
title(ws, "DATASET PENELITIAN — EVALUASI PELATIHAN CANVA",
      "Wanita Katolik RI · pelatihan 25 Agustus 2026 · pre-test 25 Agustus, post-test 25–28 Agustus 2026. "
      "Seluruh isi berkas ini berbasis ORANG, bukan sesi. Setiap angka dapat ditelusuri ke empat berkas sumber di bawah.", 6)
r = 4
r = sec(ws, r, "A. BERKAS SUMBER", 6)
r = head(ws, r, ["Berkas", "Jenis", "Dipakai untuk"], [58, 20, 74])
for a, b, c in [
    ("pretestpelatihancanva25agustus2026-...-c1bee5.xlsx", "Ekspor resmi Wayground",
     "Respons tiap peserta, waktu, metadata sesi pre-test."),
    ("post-testpelatihancanva25agustus2026-...-68e3ea.xlsx", "Ekspor resmi Wayground",
     "Respons, waktu per butir, metadata sesi post-test."),
    ("Free Printable pretest pelatihan canva 25 agustus 2026.pdf", "Naskah cetak resmi",
     "Naskah 20 butir dan keempat opsinya; pembanding untuk memverifikasi kunci pre-test."),
    ("Free Printable post-test pelatihan canva 25 agustus 2026.pdf", "Naskah cetak resmi",
     "Naskah 20 butir dan keempat opsinya; pembanding untuk memverifikasi kunci post-test."),
    ("Wayground 25 agustus 2026 canva wkri.html", "Snapshot laporan admin",
     "Matriks respons berwarna pre-test; dipakai menyaring entri jawaban semu pada ekspor XLSX."),
]:
    r = row(ws, r, [a, b, c], h=24, fs=9)
r += 1

r = sec(ws, r, "B. ATURAN PENGOLAHAN — DITETAPKAN DI MUKA", 6)
r = head(ws, r, ["Keputusan", "Aturan", "Alasan"], [30, 56, 66])
for a, b, c in [
    ("Satuan analisis", "Satu baris = satu ORANG.",
     "Wayground mencatat sesi. Orang yang koneksinya terputus lalu masuk kembali tercatat sebagai dua baris."),
    ("Orang bersesi ganda", "Diambil sesi dengan butir terjawab terbanyak; bila seri, yang benarnya terbanyak.",
     "Pada keempat kasus, kedua kriteria menunjuk sesi yang sama, sehingga pilihan ini tidak mengubah hasil."),
    ("Sesi uji perangkat lunak", "Sesi atas nama Vincent pada post-test dikeluarkan seluruhnya.",
     "15 dari 20 butir dijawab dalam waktu persis 1 detik; tidak muncul pada daftar pre-test."),
    ("Butir tidak dijawab", "Dihitung TIDAK BENAR pada statistik butir, tetapi dibedakan pada matriks respons.",
     "Menyamakan kosong dengan salah pada matriks akan menghapus jejak sesi yang terputus."),
    ("Kunci jawaban", "Direkonstruksi dari pola respons, lalu diverifikasi ke naskah cetak resmi.",
     "Ekspor Wayground tidak memuat kolom kunci. Kecocokan sempurna 20/20 pada kedua tes."),
    ("Dasar uji statistik", "Hanya orang yang mengikuti KEDUA tes dan menyelesaikan post-test 20 butir.",
     "Mencampurkan sesi terputus akan mengukur kegagalan teknis, bukan pengetahuan."),
]:
    r = row(ws, r, [a, b, c], bold=(1,), h=30, fs=9)
r += 1

r = sec(ws, r, "C. ISI BERKAS", 6)
r = head(ws, r, ["Lembar", "Isi"], [30, 122])
for a, b in [
    ("01 Sensus", "Rekonsiliasi jumlah sesi menjadi jumlah orang, beserta bukti rinci tiap sesi ganda."),
    ("02 Soal Pre-test", "Naskah 20 butir dan keempat opsinya apa adanya dari naskah cetak resmi, dengan penanda kunci dan jumlah pemilih tiap opsi."),
    ("03 Soal Post-test", "Idem untuk post-test."),
    ("04 Butir Pre-test", "Statistik tiap butir: benar, salah, kosong, p, D, r-pbis."),
    ("05 Butir Post-test", "Idem untuk post-test."),
    ("06 Respons Pre-test", "Matriks 33 orang x 20 butir berisi B / S / kosong."),
    ("07 Respons Post-test", "Matriks 14 orang x 20 butir."),
    ("08 Jawaban Pre-test", "Teks opsi yang benar-benar dipilih tiap orang pada tiap butir."),
    ("09 Jawaban Post-test", "Idem untuk post-test, ditambah waktu pengerjaan per butir."),
    ("10 Berpasangan", "Delapan orang yang mengikuti kedua tes sampai tuntas, beserta seluruh uji statistik."),
    ("11 Grafik", "Seluruh grafik dengan tabel datanya di sebelah kiri."),
    ("12 Data Olah", "Format panjang siap impor ke SPSS, jamovi, atau R."),
]:
    r = row(ws, r, [a, b], bold=(1,), h=18, fs=9)
r += 1

r = sec(ws, r, "D. KETERBATASAN — BACA SEBELUM MENGUTIP", 6)
for t in [
    "1. Kedua tes memakai perangkat soal BERBEDA. Hanya 14 dari 20 konstruk beririsan dan hanya dua butir yang praktis identik. "
    "Nomor butir pre-test dan post-test tidak merujuk materi yang sama.",
    "2. Mode pelaksanaan berbeda. Pre-test dikerjakan langsung dengan batas waktu; post-test dibuka tiga hari sebagai pekerjaan rumah "
    "tanpa pengawasan, dengan materi pelatihan tersedia.",
    "3. Peserta menyusut. Dari 33 orang pre-test hanya 11 muncul kembali di post-test, dan 8 di antaranya tuntas. Yang bertahan "
    "rata-rata pre-test-nya lebih tinggi daripada kelas, sehingga terdapat bias seleksi.",
    "4. Karena ketiga hal di atas bekerja ke arah yang sama, rata-rata gain +5,00 butir diperlakukan sebagai BATAS ATAS dampak "
    "pelatihan, bukan perkiraan tak bias.",
    "5. Tidak ada kelompok pembanding, sehingga sebagian kenaikan dapat berasal dari efek mengerjakan tes serupa untuk kedua kalinya.",
    "6. Yang terukur adalah pengetahuan deklaratif. Kemampuan membuat desain memerlukan penilaian karya.",
    "7. Berkas ini memuat nama asli. Untuk publikasi yang memerlukan persetujuan etik, nama perlu diganti kode peserta.",
]:
    r = note(ws, r, t, 6, 28)
