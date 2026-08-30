# -*- coding: utf-8 -*-
"""BUKTI KEABSAHAN — 'Rekapitulasi Pre-Post Test - Canva WKRI.xlsx'

Membuktikan setiap tanda B/S pada berkas rekapitulasi berasal dari empat dokumen sumber,
melalui tiga mata rantai yang diperiksa satu per satu:

  Rantai 1  Teks jawaban tiap orang = sel pada ekspor resmi Wayground (XLSX).
  Rantai 2  Kunci tiap butir = salah satu opsi pada naskah cetak resmi (PDF), sama persis.
  Rantai 3  Tanda B/S = hasil perbandingan jawaban terhadap kunci.

Tidak ada penilaian, pembobotan, atau penyesuaian apa pun di antara ketiganya.
"""
import openpyxl, json, io, os, re
from pypdf import PdfReader

SRC = "E:/Download"
KLON = "Rekapitulasi Pre-Post Test - Canva WKRI.xlsx"
DOK = {
    "D1": ("pretestpelatihancanva25agustus2026-2026-08-25T09_22_13_634913-c1bee5.xlsx",
           "Ekspor resmi Wayground — pre-test"),
    "D2": ("post-testpelatihancanva25agustus2026-2026-08-28T14_07_30_851549-68e3ea.xlsx",
           "Ekspor resmi Wayground — post-test"),
    "D3": ("Free Printable pretest pelatihan canva 25 agustus 2026.pdf",
           "Naskah cetak resmi — pre-test"),
    "D4": ("Free Printable post-test pelatihan canva 25 agustus 2026.pdf",
           "Naskah cetak resmi — post-test"),
}

D = json.load(io.open("ds_core.json", encoding="utf-8"))
pre = json.load(io.open("core.json", encoding="utf-8"))
post = json.load(io.open("post_core.json", encoding="utf-8"))
S = json.load(io.open("stats.json", encoding="utf-8"))
QPRE = {q["no"]: q for q in pre["Q"]}
QPOST = {q["no"]: q for q in post["Q"]}
KEYPRE = {i["no"]: i["key"] for i in S["items"]}
KEYPOST = {q["no"]: q["key"] for q in post["Q"]}
CH_PRE, CH_POST = D["CH_PRE"], D["CH_POST"]
RPRE, RPOST = D["RPRE"], D["RPOST"]
PAIRT = sorted(D["PAIRT"], key=lambda x: -(x[3] - x[2]))
NAMES = [o for o, p, a, b, ta, tb in PAIRT]
PM = {o: p for o, p, a, b, ta, tb in PAIRT}

L, res = [], []
W = L.append


def nm(s):
    s = str(s)
    for a, b in (("\u2018", "'"), ("\u2019", "'"), ("\u201c", '"'), ("\u201d", '"'),
                 ("\u2013", "-"), ("\u2014", "-"), ("\u2192", "->"), ("\u2026", ""), ("\ufffd", "")):
        s = s.replace(a, b)
    return re.sub(r"\s+", " ", s).strip().lower()


def ck(t, ok, d=""):
    res.append((t, ok, d))
    W("  %-6s %-62s %s" % ("SAH" if ok else "GAGAL", t, d))


_F = [r"\s*\d{1,2}/\d{1,2}/\d{2,4},?\s*\d{1,2}:\d{2}\s*[AP]M.*$",
      r"\s*/\d{2},?\s*\d{1,2}:\d{2}\s*[AP]M.*$",
      r"\s*Free Printable.*$", r"\s*Worksheets?\s*$", r"\s+\d{1,2}/\d{1,2}\s*$"]


def cf(t):
    x = str(t)
    for p in _F:
        x = re.sub(p, "", x, flags=re.I)
    return re.sub(r"\s+", " ", x).strip()


def pdf_opts(path):
    raw = "\n".join((p.extract_text() or "") for p in PdfReader(path).pages)
    raw = re.sub(r"https://\S+", " ", raw)
    raw = re.sub(r"\b\d+/\d+\b", " ", raw)
    flat = re.sub(r"\s+", " ", raw)
    m0 = re.search(r"(?:(?<= )|^)1\.\s*[A-Z]", flat)
    flat = flat[m0.start():] if m0 else flat
    idx = [(int(m.group(1)), m.start())
           for m in re.finditer(r"(?:(?<= )|^)(\d{1,2})\.\s*(?=[A-Z])", flat)]
    idx = [(n, s) for n, s in idx if 1 <= n <= 20]
    seen, keep = set(), []
    for n, s in idx:
        if n not in seen:
            seen.add(n); keep.append((n, s))
    keep.sort(key=lambda x: x[1])
    out = {}
    for j, (n, s) in enumerate(keep):
        e = keep[j + 1][1] if j + 1 < len(keep) else len(flat)
        parts = re.split(r"\b([a-d])\)\s*", flat[s:e])
        out[n] = [(parts[k], cf(parts[k + 1])) for k in range(1, len(parts) - 1, 2)]
    return out


def export_grid(path):
    wbx = openpyxl.load_workbook(path, data_only=True)
    rows = [list(r) for r in wbx["Overview"].iter_rows(values_only=True)]
    hdr = rows[0]
    pc = {}
    for i in range(12, len(hdr)):
        if hdr[i] is None:
            continue
        m = re.match(r"^(.*?)\s*\((.*)\)\s*$", str(hdr[i]))
        pc[(m.group(1) if m else str(hdr[i])).strip()] = i
    g = {n: {} for n in pc}
    for r in rows[1:21]:
        for n, i in pc.items():
            g[n][int(r[0])] = r[i]
    return g


W("=" * 100)
W("BUKTI KEABSAHAN DAN KELENGKAPAN DATA")
W("Berkas   : %s" % KLON)
W("Format   : mengikuti berkas rujukan 'Rekapitulasi Pre-Post Test - rhe.xlsx'")
W("=" * 100)
W("")
W("DOKUMEN SUMBER")
for k, (f, d) in DOK.items():
    p = os.path.join(SRC, f)
    W("  %s  %-62s  %s" % (k, f[:62], "ADA" if os.path.isfile(p) else "TIDAK ADA"))
    W("      %s · %s byte" % (d, "{:,}".format(os.path.getsize(p)) if os.path.isfile(p) else "-"))
W("")

OPRE = pdf_opts(os.path.join(SRC, DOK["D3"][0]))
OPOST = pdf_opts(os.path.join(SRC, DOK["D4"][0]))
GPRE = export_grid(os.path.join(SRC, DOK["D1"][0]))
GPOST = export_grid(os.path.join(SRC, DOK["D2"][0]))
wb = openpyxl.load_workbook(KLON)

# ---------- RANTAI 2: kunci vs naskah PDF
W("-" * 100)
W("RANTAI 2 — KUNCI TIAP BUTIR ADA PERSIS PADA NASKAH CETAK RESMI")
W("-" * 100)
for lab, OPT, KEY, dok in (("PRE-TEST", OPRE, KEYPRE, "D3"), ("POST-TEST", OPOST, KEYPOST, "D4")):
    miss, dua = [], []
    for n in range(1, 21):
        ex = [l for l, t in OPT[n] if nm(t) == nm(KEY[n])]
        if len(ex) == 0:
            miss.append(n)
        elif len(ex) > 1:
            dua.append(n)
    ck("%s: 20 kunci cocok PERSIS pada satu opsi naskah (%s)" % (lab, dok),
       not miss and not dua, "tidak cocok: %s · ganda: %s" % (miss or "tidak ada", dua or "tidak ada"))
    nopt = sum(len(OPT[n]) for n in range(1, 21))
    ck("%s: naskah memuat 80 opsi (20 butir x 4)" % lab, nopt == 80, "%d opsi terbaca" % nopt)
W("")

# ---------- RANTAI 1 + 3
W("-" * 100)
W("RANTAI 1 DAN 3 — JAWABAN DARI EKSPOR, DAN TANDA B/S SEBAGAI HASIL PERBANDINGAN")
W("-" * 100)


def kues_rows(r0):
    ws = wb["Kuesioner"]
    out = {}
    for i in range(len(NAMES)):
        r = r0 + i
        out[ws.cell(r, 3).value] = [ws.cell(r, 4 + q).value for q in range(20)]
    return out


def revisi_rows(rh):
    ws = wb["Revisi"]
    hdr = [ws.cell(rh, 3 + j).value for j in range(len(NAMES))]
    out = {h: [] for h in hdr}
    for qi in range(20):
        r = rh + 2 + qi
        for j, h in enumerate(hdr):
            out[h].append(ws.cell(r, 3 + j).value)
    return out


for lab, r0, rh, G, CH, KEY, who, dokx, dokp in (
        ("PRE-TEST", 4, 4, GPRE, CH_PRE, KEYPRE, lambda o: PM[o], "D1", "D3"),
        ("POST-TEST", 16, 32, GPOST, CH_POST, KEYPOST, lambda o: o, "D2", "D4")):
    K = kues_rows(r0)
    R = revisi_rows(rh)
    bad_ans, bad_mark, bad_tr, n = [], [], [], 0
    for o in NAMES:
        sesi = CH[who(o)]
        for q in range(1, 21):
            n += 1
            ans = G[sesi][q]
            benar = (ans is not None) and (nm(ans) == nm(KEY[q]))
            harus = "B" if benar else "S"
            if K[o][q - 1] != harus:
                bad_mark.append((o, q))
            if R[o][q - 1] != K[o][q - 1]:
                bad_tr.append((o, q))
    ck("%s: %d tanda B/S = hasil banding jawaban (%s) terhadap kunci (%s)" % (lab, n, dokx, dokp),
       not bad_mark, "beda: %s" % (bad_mark[:3] or "tidak ada"))
    ck("%s: lembar Revisi konsisten dengan lembar Kuesioner (%d sel)" % (lab, n),
       not bad_tr, "beda: %s" % (bad_tr[:3] or "tidak ada"))
W("")

# ---------- KELENGKAPAN & FORMAT
W("-" * 100)
W("KELENGKAPAN DAN KESESUAIAN FORMAT")
W("-" * 100)
ck("Berkas memuat tepat dua lembar, sama seperti rujukan",
   wb.sheetnames == ["Kuesioner", "Revisi"], str(wb.sheetnames))
ws = wb["Kuesioner"]
ck("Kuesioner: blok Pretest dan Posttest ada",
   ws["B2"].value == "Pretest" and ws["B14"].value == "Posttest",
   "B2=%r B14=%r" % (ws["B2"].value, ws["B14"].value))
ck("Kuesioner: 20 kolom soal J1..J20",
   all(ws.cell(3, 4 + i).value == "J%d" % (i + 1) for i in range(20)), "")
ck("Kuesioner: kolom Jumlah B dan Jumlah S memakai rumus COUNTIF",
   str(ws["X4"].value).startswith("=COUNTIF") and str(ws["Y4"].value).startswith("=COUNTIF"),
   "%s" % ws["X4"].value)
ws2 = wb["Revisi"]
ck("Revisi: blok Pre-test dan Post-test pada posisi sel rujukan",
   ws2["E2"].value == "Pre-test" and ws2["E30"].value == "Post-test", "")
ck("Revisi: baris J1..J20 pada kedua blok",
   ws2["B6"].value == "J1" and ws2["B25"].value == "J20"
   and ws2["B34"].value == "J1" and ws2["B53"].value == "J20", "")
ck("Revisi: baris Jumlah B dan Jumlah S pada posisi rujukan",
   ws2["B26"].value == "Jumlah B" and ws2["B27"].value == "Jumlah S"
   and ws2["B54"].value == "Jumlah B" and ws2["B55"].value == "Jumlah S", "")
nprosa = sum(1 for w in wb for row in w.iter_rows(values_only=True)
             for v in row if v is not None and isinstance(v, str)
             and len(v) > 120 and v.count(" ") > 18)
ck("Tidak ada sel berisi paragraf naratif", nprosa == 0, "%d sel prosa" % nprosa)
ck("Tidak ada lembar tambahan di luar format rujukan", len(wb.sheetnames) == 2, "")
W("")

W("=" * 100)
ok = sum(1 for a, b, c in res if b)
W("REKAP: %d dari %d pemeriksaan SAH" % (ok, len(res)))
g = [a for a, b, c in res if not b]
if g:
    W("")
    W("BELUM SAH:")
    for a in g:
        W("   - " + a)
W("=" * 100)
W("")
W("CATATAN YANG HARUS DISERTAKAN")
W("  1. Berkas ini memuat 8 orang yang mengikuti pre-test DAN post-test sampai tuntas.")
W("     Peserta lain tidak dapat dibandingkan karena tidak mengikuti kedua tes.")
W("  2. Tanda S menggabungkan jawaban salah dan butir yang tidak dijawab, mengikuti")
W("     format rujukan yang hanya mengenal dua keadaan.")
W("  3. Nomor soal pre-test dan post-test TIDAK merujuk materi yang sama; kedua tes")
W("     memakai perangkat soal berbeda.")

io.open("klon_audit.txt", "w", encoding="utf-8").write("\n".join(L))
print("\n".join(L))
