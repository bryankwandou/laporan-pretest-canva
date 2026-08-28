# -*- coding: utf-8 -*-
import openpyxl, io, json, re
from collections import Counter

PT = "E:/Download/post-testpelatihancanva25agustus2026-2026-08-28T14_07_30_851549-68e3ea.xlsx"
wb = openpyxl.load_workbook(PT, data_only=True)

ws = wb["Overview"]
rows = [list(r) for r in ws.iter_rows(values_only=True)]
hdr = rows[0]
# participant columns start at index 12
pcols = []
for i in range(12, len(hdr)):
    if hdr[i] is None:
        continue
    nm = str(hdr[i])
    m = re.match(r"^(.*?)\s*\((.*)\)\s*$", nm)
    disp = (m.group(1) if m else nm).strip()
    pcols.append((i, disp))

QP = []
for r in rows[1:]:
    if r[0] is None:
        continue
    ans = {}
    for i, nm in pcols:
        v = r[i]
        ans[nm] = (str(v).strip() if v is not None else None)
    QP.append({
        "no": int(r[0]),
        "text": str(r[1]).strip(),
        "acc": int(str(r[4]).rstrip("%")),
        "correct": int(r[6]),
        "incorrect": int(r[9]),
        "unatt": int(r[11]),
        "avgt": str(r[5]),
        "answers": ans,
    })

# sanity: unattempted == count of None
bad = [q["no"] for q in QP if sum(1 for v in q["answers"].values() if v is None) != q["unatt"]]
print("unatt mismatch:", bad)

# key reconstruction: option whose selection count == Correct count (unique)
for q in QP:
    cnt = Counter(v for v in q["answers"].values() if v is not None)
    cands = [o for o, c in cnt.items() if c == q["correct"]]
    q["distr"] = sorted(cnt.items(), key=lambda kv: -kv[1])
    q["key"] = cands[0] if len(cands) == 1 else None
    q["key_amb"] = len(cands)
print("keys unresolved:", [(q["no"], q["key_amb"]) for q in QP if q["key"] is None])

# participants
PP = {}
for r in list(wb["Participant Data"].iter_rows(values_only=True))[1:]:
    if r[0] is None:
        continue
    nm = str(r[3]).strip()
    tt = str(r[13])
    h, m_, s = [int(x) for x in tt.split(":")]
    PP[nm] = {"name": nm, "att": int(r[4]), "acc": int(str(r[5]).rstrip("%")),
              "score": int(r[6]), "correct": int(r[7]), "incorrect": int(r[10]),
              "unatt": int(r[12]), "time_s": h * 3600 + m_ * 60 + s}

# duplicate display names in Overview: 'Yovita' and 'Yovita*'
print("overview names:", [n for _, n in pcols])
print("participant names:", list(PP))
print("QD:", [tuple(r) for r in wb["Quiz Details"].iter_rows(values_only=True)])

json.dump({"Q": QP, "P": PP, "names": [n for _, n in pcols]},
          io.open("post_core_raw.json", "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print("ok", len(QP), len(PP))
