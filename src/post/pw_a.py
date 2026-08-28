# -*- coding: utf-8 -*-
import json, io, math, statistics as st
from collections import Counter, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList

pre = json.load(io.open("core.json", encoding="utf-8"))
S = json.load(io.open("stats.json", encoding="utf-8"))
post = json.load(io.open("post_core.json", encoding="utf-8"))
C1 = json.load(io.open("cmp.json", encoding="utf-8"))
C2 = json.load(io.open("cmp2.json", encoding="utf-8"))
PD = json.load(io.open("pdf_opts.json", encoding="utf-8"))

PP, PC = pre["P"], pre["CORR"]
preit = {i["no"]: i for i in S["items"]}
QO = {q["no"]: q for q in post["Q"]}
PO, CO = post["P"], post["CORR"]
onames, pmap = post["names"], post["pmap"]
OPT, DEAD = {int(k): v for k, v in PD["OPT"].items()}, {int(k): v for k, v in PD["DEAD"].items()}

sc_post = {n: CO[n].count("C") for n in onames}
sc_pre = {n: PP[n]["correct"] for n in PP}
att = {n: 20 - CO[n].count("-") for n in onames}
COMP, INC = C2["comp"], C2["inc"]
PAIR, PAIRC = C1["PAIR"], C2["PAIRC"]
MAP = C1["MAP"]
NEWP = {int(k): v for k, v in C1["NEW_POST"].items()}
DROPP = {int(k): v for k, v in C1["DROP_PRE"].items()}
ORDP = sorted(onames, key=lambda n: (-sc_post[n], PO[n]["time_s"]))

NAVY = "12284B"; BLUE = "2F6FB5"; LGREY = "F2F5F9"
GRNH = "1E6E4E"; REDH = "A33B33"


def F(c):
    return PatternFill("solid", fgColor=c)


HDR = F(NAVY); SUB = F(BLUE)
WF = Font(color="FFFFFF", bold=True, size=10)
TH = Side(style="thin", color="D5DCE6")
BOX = Border(TH, TH, TH, TH)
GRN = F("D6ECDF"); RED = F("FADFDC"); YEL = F("FDF0D3"); GRY = F("EDEFF3")
BLU = F("E4EDF8")

wb = openpyxl.Workbook()
wb.remove(wb.active)


def head(ws, row, cols, widths=None):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row, i, c)
        cell.fill = HDR; cell.font = WF; cell.border = BOX
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[gcl(i)].width = w
    ws.row_dimensions[row].height = 34


def title(ws, text, sub="", span=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, text)
    c.fill = HDR; c.font = Font(color="FFFFFF", bold=True, size=13)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c = ws.cell(2, 1, sub)
    c.fill = F(LGREY); c.font = Font(size=9, italic=True, color="475467")
    c.alignment = Alignment(horizontal="left", indent=1, vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36


def secrow(ws, r, text, span):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    c = ws.cell(r, 1, text)
    c.fill = SUB; c.font = Font(color="FFFFFF", bold=True, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 22


def note(ws, r, t, span, h=36):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    c = ws.cell(r, 1, t)
    c.fill = F(LGREY); c.font = Font(size=10, color="333F4F")
    c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    c.border = BOX
    ws.row_dimensions[r].height = h


def putrow(ws, r, vals, ctr=(), bold=(), fills=None, h=None, fsz=10):
    for i, v in enumerate(vals, 1):
        c = ws.cell(r, i, v); c.border = BOX
        c.font = Font(size=fsz, bold=(i in bold))
        c.alignment = Alignment(horizontal="center", vertical="center") if i in ctr \
            else Alignment(wrap_text=True, vertical="top")
    if fills:
        for i, f in fills.items():
            ws.cell(r, i).fill = f
    if h:
        ws.row_dimensions[r].height = h
