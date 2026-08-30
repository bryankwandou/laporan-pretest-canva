# -*- coding: utf-8 -*-
"""Klon persis format berkas rujukan pengadilan:
'Rekapitulasi Pre-Post Test - rhe.xlsx' — dua lembar, hanya B/S, rumus COUNTIF.
Tidak ada prosa, tidak ada lembar tambahan."""
import json, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl

D = json.load(io.open("ds_core.json", encoding="utf-8"))
RPRE, RPOST = D["RPRE"], D["RPOST"]
PAIRT = sorted(D["PAIRT"], key=lambda x: -(x[3] - x[2]))

GRN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
TH = Side(style="thin", color="000000")
BOX = Border(TH, TH, TH, TH)
CEN = Alignment(horizontal="center", vertical="center")

wb = openpyxl.Workbook()
wb.remove(wb.active)

NAMES = [o for o, p, a, b, ta, tb in PAIRT]
PREOF = {o: RPRE[p] for o, p, a, b, ta, tb in PAIRT}
POSOF = {o: RPOST[o] for o, p, a, b, ta, tb in PAIRT}
N = len(NAMES)


def bs(ch):
    """Format rujukan hanya mengenal dua keadaan: B dan S."""
    return "B" if ch == "C" else "S"


def cell_bs(ws, r, c, v):
    x = ws.cell(r, c, v)
    x.font = Font(bold=True, size=10)
    x.alignment = CEN
    x.border = BOX
    x.fill = GRN if v == "B" else RED
    return x


# =============== LEMBAR 1: KUESIONER  (baris = peserta)
ws = wb.create_sheet("Kuesioner")
ws.column_dimensions["A"].width = 4.1
ws.column_dimensions["B"].width = 4.9
ws.column_dimensions["C"].width = 33.3
for i in range(4, 24):
    ws.column_dimensions[gcl(i)].width = 3.7
ws.column_dimensions["X"].width = 13.3
ws.column_dimensions["Y"].width = 13.1


def blok_kuesioner(r0, judul, getc):
    ws.cell(r0, 2, judul)
    r = r0 + 1
    ws.cell(r, 2, "No")
    ws.cell(r, 3, "Nama")
    for i in range(1, 21):
        ws.cell(r, 3 + i, "J%d" % i).alignment = CEN
    ws.cell(r, 24, "Jumlah B")
    ws.cell(r, 25, "Jumlah S")
    r += 1
    for i, nm in enumerate(NAMES, 1):
        ws.cell(r, 2, i).alignment = CEN
        ws.cell(r, 3, nm)
        cs = getc(nm)
        for qi in range(20):
            cell_bs(ws, r, 4 + qi, bs(cs[qi]))
        ws.cell(r, 24, '=COUNTIF(D%d:W%d,"B")' % (r, r))
        ws.cell(r, 25, '=COUNTIF(D%d:W%d,"S")' % (r, r))
        r += 1
    return r


r = blok_kuesioner(2, "Pretest", lambda nm: PREOF[nm])
r = blok_kuesioner(r + 2, "Posttest", lambda nm: POSOF[nm])
ws.cell(r + 1, 2, "S = salah atau tidak dijawab").font = Font(italic=True, size=9)

# =============== LEMBAR 2: REVISI  (baris = soal)
ws = wb.create_sheet("Revisi")
ws.column_dimensions["A"].width = 4.1
ws.column_dimensions["B"].width = 10.0
for i in range(3, 3 + N):
    ws.column_dimensions[gcl(i)].width = 22.0
ws.column_dimensions[gcl(3 + N)].width = 11.0
ws.column_dimensions[gcl(4 + N)].width = 11.0

KB = 3 + N          # kolom Jumlah B
KS = 4 + N          # kolom Jumlah S


def blok_revisi(rt, rh, judul, getc):
    ws.cell(rt, 5, judul)
    ws.merge_cells(start_row=rt, start_column=5, end_row=rt + 1, end_column=6)
    ws.cell(rt, 5).alignment = CEN
    ws.cell(rh, 2, "Soal")
    ws.merge_cells(start_row=rh, start_column=2, end_row=rh + 1, end_column=2)
    for j, nm in enumerate(NAMES):
        ws.cell(rh, 3 + j, nm)
        ws.merge_cells(start_row=rh, start_column=3 + j, end_row=rh + 1, end_column=3 + j)
        ws.cell(rh, 3 + j).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(rh, KB, "Jumlah B")
    ws.merge_cells(start_row=rh, start_column=KB, end_row=rh + 1, end_column=KB)
    ws.cell(rh, KS, "Jumlah S")
    ws.merge_cells(start_row=rh, start_column=KS, end_row=rh + 1, end_column=KS)
    r = rh + 2
    d0 = r
    for qi in range(20):
        ws.cell(r, 2, "J%d" % (qi + 1)).alignment = CEN
        for j, nm in enumerate(NAMES):
            cell_bs(ws, r, 3 + j, bs(getc(nm)[qi]))
        ws.cell(r, KB, '=COUNTIF(C%d:%s%d,"B")' % (r, gcl(2 + N), r))
        ws.cell(r, KS, '=COUNTIF(C%d:%s%d,"S")' % (r, gcl(2 + N), r))
        r += 1
    d1 = r - 1
    ws.cell(r, 2, "Jumlah B")
    for j in range(N):
        ws.cell(r, 3 + j, '=COUNTIF(%s%d:%s%d,"B")' % (gcl(3 + j), d0, gcl(3 + j), d1))
    r += 1
    ws.cell(r, 2, "Jumlah S")
    for j in range(N):
        ws.cell(r, 3 + j, '=COUNTIF(%s%d:%s%d,"S")' % (gcl(3 + j), d0, gcl(3 + j), d1))
    return r + 1


r = blok_revisi(2, 4, "Pre-test", lambda nm: PREOF[nm])
r = blok_revisi(r + 2, r + 4, "Post-test", lambda nm: POSOF[nm])
ws.cell(r + 1, 2, "S = salah atau tidak dijawab").font = Font(italic=True, size=9)

OUT = "Rekapitulasi Pre-Post Test - Canva WKRI.xlsx"
wb.save(OUT)
print("tersimpan:", OUT)
print("lembar:", wb.sheetnames, "| peserta:", N)
