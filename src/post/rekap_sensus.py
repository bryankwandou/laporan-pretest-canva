# -*- coding: utf-8 -*-
"""Lembar Sensus Peserta — rekonsiliasi jumlah SESI versus jumlah ORANG.
Ditambahkan ke REKAPITULASI sebagai lembar pertama."""
import json, io, re, statistics as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

pre = json.load(io.open("core.json", encoding="utf-8"))
post = json.load(io.open("post_core.json", encoding="utf-8"))
rawp = json.load(io.open("post_core_raw.json", encoding="utf-8"))
C1 = json.load(io.open("cmp.json", encoding="utf-8"))
C2 = json.load(io.open("cmp2.json", encoding="utf-8"))
PC, CO, P = pre["CORR"], post["CORR"], pre["P"]

DUPPRE = {"Sri Suyani*": "Sri Suyani", "Sri Suyani**": "Sri Suyani",
          "Aqifah*": "Aqifah", "farida johannes*": "farida johannes"}
DUPPOST = {"Yovita*": "Yovita"}

GRN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YEL = PatternFill("solid", fgColor="FFEB9C")
HDR = PatternFill("solid", fgColor="DCE6F1")
TOT = PatternFill("solid", fgColor="F2F2F2")
NAV = PatternFill("solid", fgColor="12284B")
TH = Side(style="thin", color="BFBFBF")
BOX = Border(TH, TH, TH, TH)
CEN = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(vertical="top", wrap_text=True)

wb = openpyxl.load_workbook("REKAPITULASI_PRE_POST_TEST_CANVA.xlsx")
if "Sensus Peserta" in wb.sheetnames:
    del wb["Sensus Peserta"]
ws = wb.create_sheet("Sensus Peserta", 0)

for col, w in zip("ABCDEFGH", (3, 30, 12, 12, 12, 12, 14, 60)):
    ws.column_dimensions[col].width = w


def hdr(r, cols, start=2):
    for i, t in enumerate(cols):
        c = ws.cell(r, start + i, t)
        c.font = Font(bold=True, size=10)
        c.alignment = CEN
        c.border = BOX
        c.fill = HDR
    return r + 1


def sec(r, t):
    c = ws.cell(r, 2, t)
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = NAV
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    return r + 1


def note(r, t, h=30):
    c = ws.cell(r, 2, t)
    c.font = Font(size=9, italic=True)
    c.alignment = WRAP
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.row_dimensions[r].height = h
    return r + 1


c = ws.cell(2, 2, "SENSUS PESERTA — JUMLAH SESI VERSUS JUMLAH ORANG")
c.font = Font(bold=True, size=14)
r = note(3, "Wayground mencatat SESI, bukan orang. Satu orang yang koneksinya terputus lalu masuk kembali "
            "tercatat sebagai dua baris terpisah. Lembar ini merekonsiliasi kedua cacah tersebut sehingga "
            "setiap angka pada laporan dapat ditelusuri asalnya.", 32)
r += 1

# ---------- A
r = sec(r, "A. REKONSILIASI CACAH")
r = hdr(r, ["Ukuran", "Pre-test", "Post-test", "", "", "", "Keterangan"])
uniq_pre = len({DUPPRE.get(n, n) for n in P})
uniq_post = len({DUPPOST.get(n, n) for n in post["names"]})
ROWS = [
    ("Sesi terekam di ekspor", 37, 16, "Baris apa adanya pada sheet Participant Data."),
    ("Sesi uji perangkat lunak (QA)", 0, 1, "Sesi atas nama Vincent pada post-test; 15 dari 20 butir dijawab dalam 1 detik."),
    ("Sesi peserta", 37, 15, "Setelah sesi QA dikeluarkan."),
    ("Sesi ganda dari orang yang sama", 4, 1, "Pre-test: Sri Suyani 3 sesi, Aqifah 2, farida johannes 2. Post-test: Yovita 2."),
    ("ORANG UNIK", uniq_pre, uniq_post, "INI JAWABAN atas pertanyaan 'berapa orang'."),
    ("Orang menjawab sedikitnya 1 butir", 30, 14, "Pre-test: 3 orang tidak menjawab apa pun."),
    ("Orang menyelesaikan 20 butir", 3, 10, "Pre-test dibatasi waktu 14 menit sehingga sedikit yang tuntas."),
    ("Orang mengikuti KEDUA tes", 11, 11, "Hasil pencocokan nama antar kedua daftar."),
    ("Orang mengikuti kedua tes sampai TUNTAS", 8, 8, "Dasar seluruh uji statistik pada laporan."),
]
for lab, a, b, ket in ROWS:
    ws.cell(r, 2, lab).font = Font(bold=(lab == "ORANG UNIK"), size=10)
    ws.cell(r, 2).border = BOX
    for j, v in ((0, a), (1, b)):
        cc = ws.cell(r, 3 + j, v)
        cc.alignment = CEN
        cc.border = BOX
        cc.font = Font(bold=True, size=10)
        if lab == "ORANG UNIK":
            cc.fill = GRN
        elif "ganda" in lab or "QA" in lab:
            cc.fill = YEL
    cc = ws.cell(r, 8, ket)
    cc.font = Font(size=9)
    cc.alignment = WRAP
    cc.border = BOX
    ws.row_dimensions[r].height = 26
    r += 1
r += 1

# ---------- B
r = sec(r, "B. DAFTAR SESI GANDA — BUKTI RINCI")
r = hdr(r, ["Nama pada ekspor", "Benar", "Salah", "Kosong", "Waktu (detik)", "Sesi dipakai?", "Penilaian"])
DETAIL = [("Sri Suyani", ["Sri Suyani", "Sri Suyani*", "Sri Suyani**"], PC, P),
          ("Aqifah", ["Aqifah", "Aqifah*"], PC, P),
          ("farida johannes", ["farida johannes", "farida johannes*"], PC, P)]
for orang, sesi, M, meta in DETAIL:
    best = max(sesi, key=lambda n: M[n].count("C"))
    for n in sesi:
        cs = M[n]
        pakai = (n == best)
        vals = [n, cs.count("C"), cs.count("X"), cs.count("-"), meta[n].get("time_s", 0),
                "YA" if pakai else "tidak"]
        pen = ("Sesi terlengkap; dipakai sebagai data orang ini."
               if pakai else
               ("Sesi kosong — masuk lalu keluar tanpa menjawab." if cs.count("-") == 20
                else "Sesi terputus; sebagian besar butir tidak terjawab."))
        for j, v in enumerate(vals):
            cc = ws.cell(r, 2 + j, v)
            cc.border = BOX
            cc.font = Font(size=10, bold=(j == 5 and pakai))
            if j:
                cc.alignment = CEN
        ws.cell(r, 7).fill = GRN if pakai else RED
        cc = ws.cell(r, 8, pen)
        cc.font = Font(size=9)
        cc.alignment = WRAP
        cc.border = BOX
        ws.row_dimensions[r].height = 24
        r += 1
for n in ["Yovita", "Yovita*"]:
    cs = CO[n]
    pakai = (n == "Yovita")
    for j, v in enumerate([n + "  (post-test)", cs.count("C"), cs.count("X"), cs.count("-"), "-",
                           "YA" if pakai else "tidak"]):
        cc = ws.cell(r, 2 + j, v)
        cc.border = BOX
        cc.font = Font(size=10, bold=(j == 5 and pakai))
        if j:
            cc.alignment = CEN
    ws.cell(r, 7).fill = GRN if pakai else RED
    cc = ws.cell(r, 8, "Sesi tuntas 20 butir; dipakai." if pakai else
                       "Sesi kosong — tidak satu butir pun dijawab.")
    cc.font = Font(size=9)
    cc.alignment = WRAP
    cc.border = BOX
    ws.row_dimensions[r].height = 24
    r += 1
r = note(r, "Enam dari tujuh sesi ganda praktis kosong: peserta masuk lalu keluar tanpa menjawab, "
            "atau berhenti setelah beberapa butir. Pola ini sejalan dengan kendala teknis yang sudah "
            "dicatat pada laporan pre-test, bukan dengan pengerjaan ulang untuk memperbaiki nilai.", 30)
r += 1

# ---------- C
r = sec(r, "C. DAMPAK TERHADAP ANGKA YANG DILAPORKAN")
r = hdr(r, ["Ukuran pre-test", "Basis 37 sesi", "Basis 33 orang", "Selisih", "", "", "Akibatnya"])
bestmap = {}
for n in P:
    k = DUPPRE.get(n, n)
    if k not in bestmap or PC[n].count("C") > PC[bestmap[k]].count("C"):
        bestmap[k] = n
sel = list(bestmap.values())
p37 = [sum(1 for n in P if PC[n][q] == "C") / 37 for q in range(20)]
p33 = [sum(1 for n in sel if PC[n][q] == "C") / 33 for q in range(20)]
m37 = st.mean([PC[n].count("C") for n in P])
m33 = st.mean([PC[n].count("C") for n in sel])
IMP = [("Rata-rata jawaban benar", round(m37, 2), round(m33, 2), round(m33 - m37, 2),
        "Angka 33 orang lebih tinggi karena sesi kosong tidak lagi menurunkan rata-rata."),
       ("Akurasi kelas (%)", round(m37 / 20 * 100, 1), round(m33 / 20 * 100, 1),
        round((m33 - m37) / 20 * 100, 1),
        "Basis orang menaikkan akurasi pre-test sekitar 3 poin persen."),
       ("Rata-rata tingkat kesukaran p", round(st.mean(p37), 3), round(st.mean(p33), 3),
        round(st.mean(p33) - st.mean(p37), 3),
        "Selisih terbesar pada satu butir 0,072 (Q20). Peringkat kesukaran butir tidak berubah.")]
for lab, a, b, d, ket in IMP:
    ws.cell(r, 2, lab).font = Font(size=10, bold=True)
    ws.cell(r, 2).border = BOX
    for j, v in enumerate([a, b, d]):
        cc = ws.cell(r, 3 + j, v)
        cc.alignment = CEN
        cc.border = BOX
        cc.font = Font(size=10, bold=True)
        if j == 2:
            cc.fill = YEL
    cc = ws.cell(r, 8, ket)
    cc.font = Font(size=9)
    cc.alignment = WRAP
    cc.border = BOX
    ws.row_dimensions[r].height = 26
    r += 1
r += 1

# ---------- D
r = sec(r, "D. YANG TIDAK TERDAMPAK — DAN MENGAPA")
for t in [
    "Tidak satu pun dari tiga orang bersesi ganda pada pre-test (Sri Suyani, Aqifah, farida johannes) "
    "termasuk dalam 8 peserta berpasangan. Untuk Yovita pada post-test, sesi tuntas 13 benar yang dipakai, "
    "bukan sesi kosongnya.",
    "Karena itu seluruh hasil utama laporan TIDAK berubah: rata-rata gain +5,00 butir, uji-t t(7)=3,67, "
    "Cohen dz=1,30, uji tanda p=0,0039, dan gain ternormalisasi Hake 0,440.",
    "Yang perlu dikoreksi hanyalah angka deskriptif tingkat kelas pada pre-test bila basisnya dinyatakan "
    "sebagai ORANG dan bukan SESI. Kedua basis kini dilaporkan berdampingan pada bagian C di atas.",
]:
    r = note(r, "•  " + t, 30)
r += 1

# ---------- grafik
c = ws.cell(r, 2, "Data grafik")
c.font = Font(bold=True, size=10)
r += 1
g0 = r
for i, t in enumerate(["Basis", "Pre-test", "Post-test"]):
    cc = ws.cell(r, 2 + i, t)
    cc.font = Font(bold=True, size=10)
    cc.fill = HDR
    cc.border = BOX
r += 1
for lab, a, b in [("Sesi terekam", 37, 16), ("Sesi peserta", 37, 15), ("Orang unik", uniq_pre, uniq_post),
                  ("Ikut kedua tes", 11, 11), ("Kedua tes tuntas", 8, 8)]:
    for i, v in enumerate([lab, a, b]):
        cc = ws.cell(r, 2 + i, v)
        cc.border = BOX
        cc.font = Font(size=10)
        if i:
            cc.alignment = CEN
    r += 1
g1 = r - 1
ch = BarChart()
ch.type = "col"
ch.grouping = "clustered"
ch.style = 10
ch.title = "Sesi versus orang — kedua tes"
ch.y_axis.title = "Jumlah"
ch.height, ch.width = 9, 20
ch.add_data(Reference(ws, min_col=3, max_col=4, min_row=g0, max_row=g1), titles_from_data=True)
ch.set_categories(Reference(ws, min_col=2, min_row=g0 + 1, max_row=g1))
ch.dLbls = DataLabelList()
ch.dLbls.showVal = True
ws.add_chart(ch, "J3")

wb.save("REKAPITULASI_PRE_POST_TEST_CANVA.xlsx")
print("Sensus ditambahkan. Sheet:", wb.sheetnames)
print("orang unik pre=%d post=%d" % (uniq_pre, uniq_post))
