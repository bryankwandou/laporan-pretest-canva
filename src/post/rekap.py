# -*- coding: utf-8 -*-
"""Rekapitulasi Pre-Post Test — format sederhana mengikuti berkas rujukan.
Hanya B (benar) / S (salah) / - (tidak dijawab), dua blok, dua tampilan."""
import json, io, statistics as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

pre = json.load(io.open("core.json", encoding="utf-8"))
post = json.load(io.open("post_core.json", encoding="utf-8"))
C2 = json.load(io.open("cmp2.json", encoding="utf-8"))

PC, CO = pre["CORR"], post["CORR"]
PAIRC = sorted(C2["PAIRC"], key=lambda x: -(x[3] - x[2]))

GRN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
GRY = PatternFill("solid", fgColor="E8EAED")
HDR = PatternFill("solid", fgColor="DCE6F1")
TOT = PatternFill("solid", fgColor="F2F2F2")
TH = Side(style="thin", color="BFBFBF")
BOX = Border(TH, TH, TH, TH)
CEN = Alignment(horizontal="center", vertical="center")

wb = openpyxl.Workbook()
wb.remove(wb.active)


def mark(ws, r, c, s):
    cell = ws.cell(r, c, {"C": "B", "X": "S", "-": "-"}[s])
    cell.font = Font(bold=True, size=10)
    cell.alignment = CEN
    cell.border = BOX
    cell.fill = {"C": GRN, "X": RED, "-": GRY}[s]


def colhdr(ws, r, cols, start=2):
    for i, t in enumerate(cols):
        c = ws.cell(r, start + i, t)
        c.font = Font(bold=True, size=10)
        c.alignment = CEN
        c.border = BOX
        c.fill = HDR
    return r + 1


HD = ["No", "Nama"] + ["J%d" % i for i in range(1, 21)] + ["Jumlah B", "Jumlah S"]


def widths(ws):
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 26
    for i in range(4, 24):
        ws.column_dimensions[gcl(i)].width = 4.2
    ws.column_dimensions["X"].width = 11
    ws.column_dimensions["Y"].width = 11


def roster(ws, r, label, rows):
    c = ws.cell(r, 2, label)
    c.font = Font(bold=True, size=12)
    r = colhdr(ws, r + 1, HD)
    for i, (nm, cs) in enumerate(rows, 1):
        ws.cell(r, 2, i).alignment = CEN
        ws.cell(r, 2).border = BOX
        cc = ws.cell(r, 3, nm)
        cc.border = BOX
        cc.font = Font(size=10)
        for qi, s in enumerate(cs):
            mark(ws, r, 4 + qi, s)
        for off, v in ((0, cs.count("C")), (1, cs.count("X") + cs.count("-"))):
            cc = ws.cell(r, 24 + off, v)
            cc.font = Font(bold=True, size=10)
            cc.alignment = CEN
            cc.border = BOX
            cc.fill = TOT
        r += 1
    return r + 2


# ===================== 1. KUESIONER
ws = wb.create_sheet("Kuesioner")
widths(ws)
r = roster(ws, 2, "Pretest", [(o, "".join(PC[p])) for o, p, a, b in PAIRC])
r = roster(ws, r, "Posttest", [(o, "".join(CO[o])) for o, p, a, b in PAIRC])
ws.cell(r, 2, "B = benar     S = salah     -  = tidak dijawab").font = Font(italic=True, size=9)
ws.cell(r + 1, 2, "Delapan peserta yang mengikuti pre-test dan post-test sampai tuntas. "
                  "Nama ditulis sesuai pendaftaran pada post-test.").font = Font(italic=True, size=9)

# ===================== 2. REVISI
ws = wb.create_sheet("Revisi")
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 10
for i in range(3, 11):
    ws.column_dimensions[gcl(i)].width = 22
ws.column_dimensions["K"].width = 11
ws.column_dimensions["L"].width = 11

NAMES = [o for o, p, a, b in PAIRC]
pm = {o: p for o, p, a, b in PAIRC}


def matrix(ws, r, label, getcorr):
    c = ws.cell(r, 5, label)
    c.font = Font(bold=True, size=12)
    r = colhdr(ws, r + 2, ["Soal"] + NAMES + ["Jumlah B", "Jumlah S"])
    for qi in range(20):
        c = ws.cell(r, 2, "J%d" % (qi + 1))
        c.font = Font(bold=True, size=10)
        c.alignment = CEN
        c.border = BOX
        c.fill = HDR
        nb = 0
        for j, nm in enumerate(NAMES):
            s = getcorr(nm)[qi]
            mark(ws, r, 3 + j, s)
            nb += (s == "C")
        for off, v in ((0, nb), (1, len(NAMES) - nb)):
            cc = ws.cell(r, 11 + off, v)
            cc.font = Font(bold=True, size=10)
            cc.alignment = CEN
            cc.border = BOX
            cc.fill = TOT
        r += 1
    for lab, fn in (("Jumlah B", lambda s: s.count("C")),
                    ("Jumlah S", lambda s: s.count("X") + s.count("-"))):
        c = ws.cell(r, 2, lab)
        c.font = Font(bold=True, size=10)
        c.alignment = CEN
        c.border = BOX
        c.fill = TOT
        for j, nm in enumerate(NAMES):
            cc = ws.cell(r, 3 + j, fn(getcorr(nm)))
            cc.font = Font(bold=True, size=10)
            cc.alignment = CEN
            cc.border = BOX
            cc.fill = TOT
        r += 1
    return r + 2


r = matrix(ws, 2, "Pre-test", lambda nm: "".join(PC[pm[nm]]))
r = matrix(ws, r, "Post-test", lambda nm: "".join(CO[nm]))
ws.cell(r, 2, "B = benar     S = salah     -  = tidak dijawab").font = Font(italic=True, size=9)

# ===================== 3. REKAP
ws = wb.create_sheet("Rekap")
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 5
ws.column_dimensions["C"].width = 26
for i in range(4, 9):
    ws.column_dimensions[gcl(i)].width = 12

ws.cell(2, 2, "Rekap per peserta").font = Font(bold=True, size=12)
r = colhdr(ws, 3, ["No", "Nama", "Pre-test", "Post-test", "Selisih", "Nilai pre", "Nilai post"])
p0 = r
for i, (o, p, a, b) in enumerate(PAIRC, 1):
    for j, v in enumerate([i, o, a, b, b - a, a * 5, b * 5]):
        cc = ws.cell(r, 2 + j, v)
        cc.border = BOX
        cc.font = Font(size=10)
        if j != 1:
            cc.alignment = CEN
    ws.cell(r, 6).font = Font(bold=True, size=10)
    ws.cell(r, 6).fill = GRN if b > a else (RED if b < a else GRY)
    r += 1
p1 = r - 1
ga = [a for o, p, a, b in PAIRC]
gb = [b for o, p, a, b in PAIRC]
for j, v in enumerate(["", "Rata-rata", round(st.mean(ga), 2), round(st.mean(gb), 2),
                       round(st.mean(gb) - st.mean(ga), 2),
                       round(st.mean(ga) * 5, 1), round(st.mean(gb) * 5, 1)]):
    cc = ws.cell(r, 2 + j, v)
    cc.border = BOX
    cc.fill = TOT
    cc.font = Font(bold=True, size=10)
    cc.alignment = CEN
rr = r + 2

ch = BarChart()
ch.type = "col"
ch.grouping = "clustered"
ch.style = 10
ch.title = "Jumlah jawaban benar: pre-test vs post-test"
ch.y_axis.title = "Jawaban benar (dari 20)"
ch.height, ch.width = 9, 20
ch.add_data(Reference(ws, min_col=4, max_col=5, min_row=p0 - 1, max_row=p1), titles_from_data=True)
ch.set_categories(Reference(ws, min_col=3, min_row=p0, max_row=p1))
ch.dLbls = DataLabelList()
ch.dLbls.showVal = True
ws.add_chart(ch, "J3")

ws.cell(rr, 2, "Rekap per soal").font = Font(bold=True, size=12)
r = colhdr(ws, rr + 1, ["Soal", "Benar pre", "Benar post", "Selisih"])
q0 = r
for qi in range(20):
    nbp = sum(1 for o, p, a, b in PAIRC if PC[p][qi] == "C")
    nbq = sum(1 for o, p, a, b in PAIRC if CO[o][qi] == "C")
    for j, v in enumerate(["J%d" % (qi + 1), nbp, nbq, nbq - nbp]):
        cc = ws.cell(r, 2 + j, v)
        cc.border = BOX
        cc.alignment = CEN
        cc.font = Font(size=10)
    ws.cell(r, 5).font = Font(bold=True, size=10)
    ws.cell(r, 5).fill = GRN if nbq > nbp else (RED if nbq < nbp else GRY)
    r += 1
q1 = r - 1

ch2 = BarChart()
ch2.type = "col"
ch2.grouping = "clustered"
ch2.style = 12
ch2.title = "Jumlah peserta menjawab benar per soal (8 peserta)"
ch2.y_axis.title = "Jumlah peserta benar"
ch2.height, ch2.width = 9, 22
ch2.add_data(Reference(ws, min_col=3, max_col=4, min_row=q0 - 1, max_row=q1), titles_from_data=True)
ch2.set_categories(Reference(ws, min_col=2, min_row=q0, max_row=q1))
ws.add_chart(ch2, "J%d" % (rr + 2))
ws.cell(r + 1, 2, "Nilai = jumlah benar x 5, sehingga 20 butir setara nilai 100.").font = Font(italic=True, size=9)
ws.cell(r + 2, 2, "Nomor soal pre-test dan post-test TIDAK merujuk materi yang sama; "
                  "kedua tes memakai perangkat soal berbeda.").font = Font(italic=True, size=9)

# ===================== 4. ROSTER LENGKAP
ws = wb.create_sheet("Kuesioner Lengkap")
widths(ws)
pn = sorted(pre["P"], key=lambda n: -PC[n].count("C"))
on = sorted(post["names"], key=lambda n: -CO[n].count("C"))
r = roster(ws, 2, "Pretest — seluruh 37 sesi", [(n, "".join(PC[n])) for n in pn])
r = roster(ws, r, "Posttest — seluruh 15 sesi (sesi QA tester dikeluarkan)",
           [(n, "".join(CO[n])) for n in on])
ws.cell(r, 2, "B = benar     S = salah     -  = tidak dijawab").font = Font(italic=True, size=9)
ws.cell(r + 1, 2, "Kolom Jumlah S menggabungkan jawaban salah dan butir yang tidak dijawab, "
                  "mengikuti format berkas rujukan.").font = Font(italic=True, size=9)

wb.save("REKAPITULASI_PRE_POST_TEST_CANVA.xlsx")
print("tersimpan: %d sheet, %d grafik" % (len(wb.sheetnames), sum(len(w._charts) for w in wb)))
print(wb.sheetnames)
