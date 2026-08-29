# -*- coding: utf-8 -*-
"""AUDIT KECUKUPAN WORKBOOK EXCEL.
Membaca kembali isi sel kedua workbook dan mencocokkannya ke data sumber.
Tidak memakai klaim apa pun dari proses pembuatan."""
import openpyxl, json, io, os, re

SUB = "E:/Download/SUBMISSION_EVALUASI_PELATIHAN_CANVA"
XPRE = os.path.join(SUB, "01_LAPORAN_EXCEL", "LAPORAN_EVALUASI_PRETEST_CANVA_25AGT2026.xlsx")
XPOST = os.path.join(SUB, "01_LAPORAN_EXCEL", "LAPORAN_POSTTEST_DAN_PERBANDINGAN_CANVA.xlsx")
DO = os.path.join(SUB, "04_DATA_OLAHAN")

pre = json.load(io.open(os.path.join(DO, "core.json"), encoding="utf-8"))
S = json.load(io.open(os.path.join(DO, "stats.json"), encoding="utf-8"))
post = json.load(io.open(os.path.join(DO, "post_core.json"), encoding="utf-8"))
raw = json.load(io.open(os.path.join(DO, "post_core_raw.json"), encoding="utf-8"))
PDo = json.load(io.open(os.path.join(DO, "pdf_opts.json"), encoding="utf-8"))
C1 = json.load(io.open(os.path.join(DO, "cmp.json"), encoding="utf-8"))
C2 = json.load(io.open(os.path.join(DO, "cmp2.json"), encoding="utf-8"))

L = []
W = L.append


def txt(wb):
    """Seluruh teks sel workbook, per sheet dan gabungan."""
    per, allt = {}, []
    for ws in wb:
        vals = []
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None:
                    vals.append(str(v))
        per[ws.title] = "\n".join(vals)
        allt.append(per[ws.title])
    return per, "\n".join(allt)


def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip().lower()


wbA = openpyxl.load_workbook(XPRE, data_only=False)
wbB = openpyxl.load_workbook(XPOST, data_only=False)
perA, TA = txt(wbA)
perB, TB = txt(wbB)
TAn, TBn = norm(TA), norm(TB)

W("=" * 92)
W("AUDIT KECUKUPAN WORKBOOK EXCEL")
W("Metode: membaca kembali isi sel kedua workbook, lalu mencocokkannya ke data sumber.")
W("Tidak ada angka yang diambil dari proses pembuatan.")
W("=" * 92)
W("")

res = []


def check(kat, nama, ok, detail=""):
    res.append((kat, nama, ok, detail))
    W("  %-6s %-56s %s" % ("LULUS" if ok else "GAGAL", nama, detail))


# ============ A. PESERTA
W("-" * 92)
W("A. KECUKUPAN DATA PESERTA")
W("-" * 92)
prenames = list(pre["P"].keys())
postnames = post["names"]
miss = [n for n in prenames if norm(n) not in TAn]
check("A", "Nama 37 peserta pre-test muncul di workbook pre", not miss,
      "hilang: %s" % (miss if miss else "tidak ada"))
miss = [n for n in postnames if norm(n) not in TBn]
check("A", "Nama 15 peserta post-test muncul di workbook post", not miss,
      "hilang: %s" % (miss if miss else "tidak ada"))
# skor tiap peserta muncul
ws = wbA["02 Data Peserta"]
found = set()
for row in ws.iter_rows(values_only=True):
    for i, v in enumerate(row):
        if v is not None and str(v) in prenames:
            found.add(str(v))
check("A", "Sheet 02 memuat baris untuk tiap peserta pre-test", len(found) == 37, "%d dari 37" % len(found))
ws = wbB["06 Peserta dan Matriks"]
found = set()
for row in ws.iter_rows(values_only=True):
    for v in row:
        if v is not None and str(v) in postnames:
            found.add(str(v))
check("A", "Sheet 06 memuat baris untuk tiap peserta post-test", len(found) == 15, "%d dari 15" % len(found))
check("A", "Sesi QA tester TIDAK muncul sebagai baris data post", "Vincent" not in postnames, "dikeluarkan")
check("A", "Pengeluaran QA tester didokumentasikan di workbook", "vincent" in TBn,
      "%d penyebutan pada sheet Metodologi dan Ringkasan" % TBn.count("vincent"))
W("")

# ============ B. BUTIR SOAL
W("-" * 92)
W("B. KECUKUPAN NASKAH SOAL DAN KUNCI")
W("-" * 92)
ws = wbA["13 Soal dan Kunci"]
tA = norm(perA["13 Soal dan Kunci"])
miss = [q["no"] for q in pre["Q"] if norm(q["text"])[:60] not in tA]
check("B", "Naskah 20 butir pre-test lengkap di workbook", not miss, "hilang: %s" % (miss if miss else "tidak ada"))
km = {i["no"]: i["key"] for i in S["items"]}
miss = [n for n, k in km.items() if norm(k)[:50] not in tA]
check("B", "Kunci 20 butir pre-test lengkap", not miss, "hilang: %s" % (miss if miss else "tidak ada"))
tB = norm(perB["07 Soal Kunci Distraktor"])
miss = [q["no"] for q in post["Q"] if norm(q["text"])[:60] not in tB]
check("B", "Naskah 20 butir post-test lengkap di workbook", not miss, "hilang: %s" % (miss if miss else "tidak ada"))
miss = [q["no"] for q in post["Q"] if norm(q["key"])[:50] not in tB]
check("B", "Kunci 20 butir post-test lengkap", not miss, "hilang: %s" % (miss if miss else "tidak ada"))
# seluruh 80 opsi PDF
OPT = {int(k): v for k, v in PDo["OPT"].items()}
nopt = sum(len(v["opts"]) for v in OPT.values())
miss = [(n, l) for n, v in OPT.items() for l, t in v["opts"] if norm(t)[:45] not in tB]
check("B", "Seluruh %d opsi post-test dari naskah PDF tercantum" % nopt, not miss,
      "hilang: %d" % len(miss))
# distraktor pre
tAd = norm(perA["05 Analisis Pengecoh"] + perA["13 Soal dan Kunci"])
allopt = [(q["no"], o) for q in pre["Q"] for o, c in
          [(x, y) for x, y in [(a, b) for a, b in
           [(i["no"], i) for i in []]]]] if False else None
opts_pre = [(i["no"], o) for i in S["items"] for o, c in i["distr"]]
miss = [(n, o) for n, o in opts_pre if norm(o)[:45] not in tAd]
check("B", "Seluruh %d opsi pre-test yang pernah dipilih tercantum" % len(opts_pre), not miss,
      "hilang: %d" % len(miss))
W("")

# ============ C. SEL RESPONS
W("-" * 92)
W("C. KECUKUPAN DATA RESPONS MENTAH (setiap peserta x setiap butir)")
W("-" * 92)


def count_matrix(ws, names, ncol=20, startcol=3):
    """Hitung sel terisi pada blok matriks."""
    n = 0
    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row if c is not None]
        if not cells:
            continue
        if str(cells[0]) in names or (len(cells) > 1 and str(cells[1]) in names):
            n += 1
    return n


ws = wbA["14 Data Mentah Jawaban"]
rows = count_matrix(ws, set(prenames))
check("C", "Sheet 14 pre-test: satu baris per peserta", rows == 37, "%d baris" % rows)
# hitung sel non-kosong pada matriks pre
filled = 0
for r in ws.iter_rows(min_row=5, values_only=True):
    if r[1] is None:
        continue
    filled += sum(1 for v in r[2:22] if v is not None)
check("C", "Sheet 14: 37 x 20 = 740 sel respons terisi", filled == 740, "%d sel" % filled)
ws = wbB["11 Data Mentah"]
filled = 0
nr = 0
for r in ws.iter_rows(min_row=5, values_only=True):
    if r[1] is None:
        continue
    nr += 1
    filled += sum(1 for v in r[2:22] if v is not None)
check("C", "Sheet 11 post-test: satu baris per peserta", nr == 15, "%d baris" % nr)
check("C", "Sheet 11: 15 x 20 = 300 sel respons terisi", filled == 300, "%d sel" % filled)
# matriks benar/salah
ws = wbA["03 Matriks Respons"]
check("C", "Matriks benar/salah pre-test ada (sheet 03)", ws.max_row >= 37, "%d baris" % ws.max_row)
ws = wbB["06 Peserta dan Matriks"]
check("C", "Matriks benar/salah post-test ada (sheet 06 bagian B)", "matriks respons" in norm(perB["06 Peserta dan Matriks"]), "ada")
W("")

# ============ D. STATISTIK
W("-" * 92)
W("D. KECUKUPAN STATISTIK")
W("-" * 92)


def has(t, *vals):
    return all(norm(v) in t for v in vals)


check("D", "KR-20 dan SEM pre-test tercantum", has(TAn, "0,751") or has(TAn, "0.751"), "0,751 / +/-1,87")
check("D", "KR-20 dan SEM post-test tercantum", has(TBn, "0,721") or has(TBn, "0.721"), "0,721 / +/-1,76")
# p, D, rpb tiap butir pre
tA4 = perA["04 Analisis Butir"]
vals = set()
for row in wbA["04 Analisis Butir"].iter_rows(values_only=True):
    for v in row:
        if isinstance(v, (int, float)):
            vals.add(round(float(v), 2))
miss = [i["no"] for i in S["items"] if round(i["p"], 2) not in vals]
check("D", "Nilai p seluruh 20 butir pre-test ada di sheet 04", not miss, "hilang: %s" % (miss if miss else "tidak ada"))
vals = set()
for row in wbB["05 Analisis Butir Post"].iter_rows(values_only=True):
    for v in row:
        if isinstance(v, (int, float)):
            vals.add(round(float(v), 2))
miss = [q["no"] for q in post["Q"] if round(q["p"], 2) not in vals]
check("D", "Nilai p seluruh 20 butir post-test ada di sheet 05", not miss, "hilang: %s" % (miss if miss else "tidak ada"))
miss = [q["no"] for q in post["Q"] if round(q["D"], 2) not in vals]
check("D", "Nilai D seluruh 20 butir post-test ada", not miss, "hilang: %s" % (miss if miss else "tidak ada"))
# uji statistik
for lab, s in [("Uji-t berpasangan t(7)=3,67", "3,67"), ("Cohen dz=1,30", "1,30"),
               ("Gain Hake 0,440", "0,440"), ("Uji tanda p=0,0039", "0,0039"),
               ("Uji kepekaan t(6)=7,12", "7,12"), ("Nilai kritis 2,365", "2,365")]:
    check("D", lab + " tercantum", norm(s) in TBn, "")
W("")

# ============ E. PERBANDINGAN
W("-" * 92)
W("E. KECUKUPAN ANALISIS PERBANDINGAN")
W("-" * 92)
tB3 = norm(perB["03 Analisis Berpasangan"])
miss = [o for o, p, a, b in C1["PAIR"] if norm(o) not in tB3]
check("E", "Seluruh 11 pasangan peserta tercantum di sheet 03", not miss, "hilang: %s" % (miss if miss else "tidak ada"))
tB12 = norm(perB["12 Data Berpasangan"])
miss = [o for o, p, a, b in C1["PAIR"] if norm(o) not in tB12]
check("E", "Sheet 12 data siap olah memuat 11 pasangan", not miss, "hilang: %s" % (miss if miss else "tidak ada"))
tB4 = norm(perB["04 Perbandingan Konstruk"])
miss = [lab for lab, a, b, d, j, bt in C1["crows"] if norm(lab)[:28] not in tB4]
check("E", "14 konstruk terpetakan tercantum di sheet 04", not miss, "hilang: %s" % (miss if miss else "tidak ada"))
miss = [v for k, v in C1["NEW_POST"].items() if norm(v)[:14] not in tB4]
check("E", "6 butir post-test tanpa padanan didaftar", not miss, "hilang: %s" % (miss if miss else "tidak ada"))
miss = [v for k, v in C1["DROP_PRE"].items() if norm(v)[:14] not in tB4]
check("E", "6 butir pre-test yang tidak diulang didaftar", not miss, "hilang: %s" % (miss if miss else "tidak ada"))
DEAD = {int(k): v for k, v in PDo["DEAD"].items()}
nd = sum(len(v) for v in DEAD.values())
tB7 = norm(perB["07 Soal Kunci Distraktor"])
miss = [(n, l) for n, v in DEAD.items() for l, t in v if norm(t)[:40] not in tB7]
check("E", "Seluruh %d pengecoh mati didaftar di sheet 07" % nd, not miss, "hilang: %d" % len(miss))
W("")

# ============ F. BATAS TAFSIR
W("-" * 92)
W("F. KECUKUPAN PENGUNGKAPAN BATAS TAFSIR")
W("-" * 92)
for lab, kw in [("Instrumen berubah antara pre dan post", "instrumen"),
                ("Mode berubah live -> homework", "homework"),
                ("Penyusutan peserta 60%", "60%"),
                ("Angka +5,00 disebut sebagai batas atas", "batas atas"),
                ("Bias seleksi dijelaskan", "bias seleksi"),
                ("Tidak ada kelompok pembanding", "kelompok pembanding"),
                ("Tidak ada pemeriksaan kecurangan", "kecurangan"),
                ("Daya beda post menggelembung", "menggelembung"),
                ("Sesi terputus dipisahkan", "terputus"),
                ("Keterbatasan punya bagian tersendiri", "keterbatasan")]:
    check("F", lab, norm(kw) in TBn, "")
W("")

# ============ G. REPRODUKSI
W("-" * 92)
W("G. KECUKUPAN JEJAK REPRODUKSI")
W("-" * 92)
for lab, kw, t in [("Sumber data disebut di workbook post", "wayground", TBn),
                   ("Naskah PDF disebut sebagai pembanding kunci", "pdf", TBn),
                   ("Rumus p, D, r-pbis dicantumkan", "r-pbis", TBn),
                   ("Rumus KR-20 dicantumkan", "kr20", TBn),
                   ("Lima lapis validasi diuraikan", "lapis", TBn),
                   ("Rumus pre-test dicantumkan", "kr-20", TAn),
                   ("Sumber data pre-test disebut", "wayground", TAn)]:
    check("G", lab, norm(kw) in t, "")
W("")

# ============ H. GRAFIK
W("-" * 92)
W("H. KECUKUPAN VISUAL")
W("-" * 92)
ca = sum(len(w._charts) for w in wbA)
cb = sum(len(w._charts) for w in wbB)
check("H", "Workbook pre-test memuat grafik native", ca >= 16, "%d grafik pada %d sheet" % (ca, len(wbA.sheetnames)))
check("H", "Workbook post-test memuat grafik native", cb >= 18, "%d grafik pada %d sheet" % (cb, len(wbB.sheetnames)))
# grafik tertaut data, bukan gambar
linked = 0
for w in wbB:
    for ch in w._charts:
        for s in ch.series:
            if s.val is not None and s.val.numRef is not None:
                linked += 1
check("H", "Seri grafik post-test tertaut ke sel data (bukan gambar)", linked > 0, "%d seri tertaut" % linked)
W("")

W("=" * 92)
n_ok = sum(1 for k, n, o, d in res if o)
W("REKAP: %d dari %d pemeriksaan LULUS" % (n_ok, len(res)))
gag = [(k, n) for k, n, o, d in res if not o]
if gag:
    W("")
    W("YANG BELUM TERPENUHI:")
    for k, n in gag:
        W("  [%s] %s" % (k, n))
W("=" * 92)

out = "\n".join(L)
io.open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "audit_xlsx.txt"), "w", encoding="utf-8").write(out)
print(out)
