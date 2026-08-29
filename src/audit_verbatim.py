# -*- coding: utf-8 -*-
"""AUDIT VERBATIM.
Menguji apakah SELURUH naskah soal, opsi, jawaban mentah, dan kolom statistik
dari ekspor resmi Wayground benar-benar tersalin UTUH ke dalam sel Excel.
Pencocokan dilakukan atas seluruh teks, bukan awalannya.
"""
import openpyxl, json, io, os, re, sys

SUB = "E:/Download/SUBMISSION_EVALUASI_PELATIHAN_CANVA"
XPRE = os.path.join(SUB, "01_LAPORAN_EXCEL", "LAPORAN_EVALUASI_PRETEST_CANVA_25AGT2026.xlsx")
XPOST = os.path.join(SUB, "01_LAPORAN_EXCEL", "LAPORAN_POSTTEST_DAN_PERBANDINGAN_CANVA.xlsx")
SRC = os.path.join(SUB, "03_DATA_SUMBER")
SPRE = os.path.join(SRC, "pretestpelatihancanva25agustus2026-2026-08-25T09_22_13_634913-c1bee5.xlsx")
SPOST = os.path.join(SRC, "post-testpelatihancanva25agustus2026-2026-08-28T14_07_30_851549-68e3ea.xlsx")
DO = os.path.join(SUB, "04_DATA_OLAHAN")

pre = json.load(io.open(os.path.join(DO, "core.json"), encoding="utf-8"))
S = json.load(io.open(os.path.join(DO, "stats.json"), encoding="utf-8"))
post = json.load(io.open(os.path.join(DO, "post_core.json"), encoding="utf-8"))
PDo = json.load(io.open(os.path.join(DO, "pdf_opts.json"), encoding="utf-8"))
TJ = json.load(io.open(os.path.join(DO, "post_time.json"), encoding="utf-8"))

L = []
W = L.append
res = []


def norm(s):
    """Normalisasi ringan: rapikan spasi dan tanda kutip, huruf kecil.
    TIDAK memotong teks — seluruh isi tetap dibandingkan."""
    s = str(s)
    s = s.replace("\u2018", "'").replace("\u2019", "'")
    s = s.replace("\u201c", '"').replace("\u201d", '"')
    s = s.replace("\u2013", "-").replace("\u2014", "-").replace("\u2192", "->")
    s = s.replace("\ufffd", "").replace("\u2026", "")
    s = re.sub(r"\s+", " ", s)
    return s.strip().lower()


def cellbag(wb):
    """Kumpulan seluruh isi sel (dinormalisasi) + teks gabungan."""
    bag, parts = set(), []
    for ws in wb:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is None:
                    continue
                n = norm(v)
                if n:
                    bag.add(n)
                    parts.append(n)
    return bag, "\n".join(parts)


def check(nama, ok, detail=""):
    res.append((nama, ok, detail))
    W("  %-6s %-62s %s" % ("LULUS" if ok else "GAGAL", nama, detail))


W("=" * 100)
W("AUDIT VERBATIM — KELENGKAPAN SALINAN NASKAH SOAL DAN STATISTIK MENTAH KE EXCEL")
W("Pencocokan atas SELURUH teks (bukan awalan). Sumber pembanding: ekspor resmi Wayground.")
W("=" * 100)
W("")

wbA = openpyxl.load_workbook(XPRE, data_only=False)
wbB = openpyxl.load_workbook(XPOST, data_only=False)
bagA, TXA = cellbag(wbA)
bagB, TXB = cellbag(wbB)


def inwb(t, bag, blob):
    """Cocok bila teks utuh menjadi isi satu sel, atau termuat utuh dalam satu sel."""
    n = norm(t)
    return n in bag or n in blob


# ================= 1. NASKAH SOAL UTUH =================
W("-" * 100)
W("1. NASKAH SOAL — SELURUH KALIMAT, BUKAN POTONGAN")
W("-" * 100)
miss = [q["no"] for q in pre["Q"] if not inwb(q["text"], bagA, TXA)]
check("Naskah 20 butir PRE-TEST tersalin utuh", not miss, "gagal pada butir: %s" % (miss or "tidak ada"))
tot = sum(len(norm(q["text"])) for q in pre["Q"])
W("         total %d karakter naskah pre-test diperiksa" % tot)
miss = [q["no"] for q in post["Q"] if not inwb(q["text"], bagB, TXB)]
check("Naskah 20 butir POST-TEST tersalin utuh", not miss, "gagal pada butir: %s" % (miss or "tidak ada"))
tot = sum(len(norm(q["text"])) for q in post["Q"])
W("         total %d karakter naskah post-test diperiksa" % tot)
W("")

# ================= 2. OPSI DAN KUNCI UTUH =================
W("-" * 100)
W("2. SELURUH OPSI JAWABAN DAN KUNCI — UTUH")
W("-" * 100)
OPT = {int(k): v for k, v in PDo["OPT"].items()}
allopt = [(n, l, t) for n, v in OPT.items() for l, t in v["opts"]]
miss = [(n, l) for n, l, t in allopt if not inwb(t, bagB, TXB)]
check("Seluruh %d opsi POST-TEST dari naskah PDF resmi" % len(allopt), not miss,
      "gagal: %s" % (miss or "tidak ada"))
km = {i["no"]: i["key"] for i in S["items"]}
miss = [n for n, k in km.items() if not inwb(k, bagA, TXA)]
check("Kunci 20 butir PRE-TEST utuh", not miss, "gagal: %s" % (miss or "tidak ada"))
miss = [q["no"] for q in post["Q"] if not inwb(q["key"], bagB, TXB)]
check("Kunci 20 butir POST-TEST utuh", not miss, "gagal: %s" % (miss or "tidak ada"))
optpre = sorted({(i["no"], o) for i in S["items"] for o, c in i["distr"]})
miss = [(n, o[:30]) for n, o in optpre if not inwb(o, bagA, TXA)]
check("Seluruh %d opsi PRE-TEST yang pernah dipilih" % len(optpre), not miss,
      "gagal: %d" % len(miss))
W("")

# ================= 3. JAWABAN MENTAH TIAP SEL =================
W("-" * 100)
W("3. TEKS JAWABAN MENTAH — SETIAP PESERTA x SETIAP BUTIR, DIBANDINGKAN KE EKSPOR RESMI")
W("-" * 100)


def sheet_cells(wb, name):
    ws = wb[name]
    return [[v for v in row] for row in ws.iter_rows(values_only=True)]


# --- POST: bandingkan sel per sel ke Overview ekspor resmi
wsrc = openpyxl.load_workbook(SPOST, data_only=True)
orows = [list(r) for r in wsrc["Overview"].iter_rows(values_only=True)]
ohdr = orows[0]
pcol = {}
for i in range(12, len(ohdr)):
    if ohdr[i] is None:
        continue
    m = re.match(r"^(.*?)\s*\((.*)\)\s*$", str(ohdr[i]))
    pcol[(m.group(1) if m else str(ohdr[i])).strip()] = i

wsB = wbB["11 Data Mentah"]
grid = {}
hdr = None
for row in wsB.iter_rows(values_only=True):
    if row[1] == "Nama":
        hdr = row
        continue
    if hdr and row[1] is not None and str(row[1]) in pcol:
        grid[str(row[1])] = list(row[2:22])

bad, checked = [], 0
for name, ci in pcol.items():
    if name == "Vincent":
        continue
    for qi, r in enumerate(orows[1:21]):
        src = r[ci]
        got = grid.get(name, [None] * 20)[qi]
        checked += 1
        if src is None:
            if norm(got) not in ("—", "-", ""):
                bad.append((name, qi + 1, "seharusnya kosong"))
        else:
            if norm(got) != norm(src):
                bad.append((name, qi + 1, "beda teks"))
check("POST-TEST: %d sel jawaban identik dengan ekspor resmi" % checked, not bad,
      "beda: %s" % (bad[:4] if bad else "tidak ada"))

# --- PRE: bandingkan ke answers_real pada core.json (hasil penyaringan entri hantu)
wsA = wbA["14 Data Mentah Jawaban"]
gridA = {}
for row in wsA.iter_rows(min_row=5, values_only=True):
    if row[1] is None:
        continue
    gridA[str(row[1])] = list(row[2:22])
badA, checkedA = [], 0
for n in pre["P"]:
    for qi, q in enumerate(pre["Q"]):
        src = q.get("answers_real", {}).get(n)
        got = gridA.get(n, [None] * 20)[qi]
        checkedA += 1
        if src is None:
            if norm(got) not in ("—", "-", ""):
                badA.append((n, qi + 1))
        elif norm(got) != norm(src):
            badA.append((n, qi + 1))
check("PRE-TEST: %d sel jawaban identik dengan data terverifikasi" % checkedA, not badA,
      "beda: %s" % (badA[:4] if badA else "tidak ada"))
W("")

# ================= 4. KOLOM STATISTIK MENTAH — OVERVIEW =================
W("-" * 100)
W("4. KOLOM STATISTIK MENTAH PER BUTIR (sheet Overview ekspor resmi)")
W("-" * 100)


def numbag(wb, sheets=None):
    s = set()
    for ws in wb:
        if sheets and ws.title not in sheets:
            continue
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    s.add(round(float(v), 4))
    return s


nB = numbag(wbB)
nA = numbag(wbA)
# post: Correct / Incorrect / Unattempted / Accuracy / Avg time per butir
colmap = [("Correct", 6), ("Incorrect", 9), ("Unattempted", 11)]
for lab, ci in colmap:
    vals = [(int(r[0]), r[ci]) for r in orows[1:21]]
    W("         %s per butir dari ekspor: %s" % (lab, [v for _, v in vals]))
# nilai yang sudah dihitung ulang tanpa Vincent tersedia di workbook
miss = [q["no"] for q in post["Q"] if float(q["correct_excl"]) not in nB]
check("POST: jumlah BENAR per butir (hitung ulang tanpa QA) ada di workbook", not miss,
      "gagal: %s" % (miss or "tidak ada"))
miss = [q["no"] for q in post["Q"] if float(q["incorrect_excl"]) not in nB]
check("POST: jumlah SALAH per butir ada di workbook", not miss, "gagal: %s" % (miss or "tidak ada"))
miss = [q["no"] for q in post["Q"] if float(q["unatt_excl"]) not in nB]
check("POST: jumlah TIDAK DIJAWAB per butir ada di workbook", not miss, "gagal: %s" % (miss or "tidak ada"))
miss = [q["no"] for q in post["Q"] if round(q["p"], 2) not in nB]
check("POST: tingkat kesukaran p tiap butir", not miss, "gagal: %s" % (miss or "tidak ada"))
miss = [q["no"] for q in post["Q"] if round(q["D"], 2) not in nB]
check("POST: daya beda D tiap butir", not miss, "gagal: %s" % (miss or "tidak ada"))
miss = [q["no"] for q in post["Q"] if round(q["rpb"], 2) not in nB]
check("POST: korelasi r-pbis tiap butir", not miss, "gagal: %s" % (miss or "tidak ada"))
# waktu rata-rata Wayground per butir (detik)
AVG = {int(k): v for k, v in TJ["AVG"].items()}
miss = [q for q, v in AVG.items() if v is not None and float(v) not in nB]
check("POST: waktu rata-rata Wayground tiap butir (detik)", not miss, "gagal: %s" % (miss or "tidak ada"))
# pre
for lab, key, rnd in [("PRE: jumlah BENAR per butir", "correct", 0),
                      ("PRE: jumlah SALAH per butir", "incorrect", 0)]:
    miss = [q["no"] for q in pre["Q"] if float(q[key]) not in nA]
    check(lab, not miss, "gagal: %s" % (miss or "tidak ada"))
miss = [i["no"] for i in S["items"] if round(i["p"], 2) not in nA]
check("PRE: tingkat kesukaran p tiap butir", not miss, "gagal: %s" % (miss or "tidak ada"))
miss = [i["no"] for i in S["items"] if round(i["D"], 2) not in nA]
check("PRE: daya beda D tiap butir", not miss, "gagal: %s" % (miss or "tidak ada"))
miss = [i["no"] for i in S["items"] if round(i["rpb"], 3) not in nA and round(i["rpb"], 2) not in nA]
check("PRE: korelasi r-pbis tiap butir", not miss, "gagal: %s" % (miss or "tidak ada"))
W("")

# ================= 5. KOLOM STATISTIK MENTAH — PARTICIPANT DATA =================
W("-" * 100)
W("5. KOLOM STATISTIK MENTAH PER PESERTA (sheet Participant Data ekspor resmi)")
W("-" * 100)
prow = [list(r) for r in wsrc["Participant Data"].iter_rows(values_only=True)][1:]
PDs = {}
for r in prow:
    if r[0] is None:
        continue
    PDs[str(r[3]).strip()] = {"att": r[4], "acc": r[5], "score": r[6], "correct": r[7],
                              "incorrect": r[10], "unatt": r[12], "time": r[13]}
for lab, k, cast in [("jumlah butir DIJAWAB", "att", int),
                     ("jumlah BENAR", "correct", int),
                     ("jumlah SALAH", "incorrect", int),
                     ("jumlah TIDAK DIJAWAB", "unatt", int)]:
    miss = [n for n in post["names"] if cast(PDs[n][k]) not in nB]
    check("POST: %s tiap peserta" % lab, not miss, "gagal: %s" % (miss or "tidak ada"))


def tsec(t):
    h, m, s = [int(x) for x in str(t).split(":")]
    return h * 3600 + m * 60 + s


miss = [n for n in post["names"] if float(tsec(PDs[n]["time"])) not in nB]
check("POST: waktu total tiap peserta (detik)", not miss, "gagal: %s" % (miss or "tidak ada"))
# akurasi peserta
miss = [n for n in post["names"] if float(str(PDs[n]["acc"]).rstrip("%")) not in nB]
check("POST: akurasi persen tiap peserta", not miss, "gagal: %s" % (miss or "tidak ada"))
# skor poin Wayground
missS = [n for n in post["names"] if float(PDs[n]["score"]) not in nB]
check("POST: skor poin Wayground tiap peserta", not missS,
      "TIDAK tersalin untuk %d peserta" % len(missS) if missS else "tidak ada")
# PRE participant
wsrcA = openpyxl.load_workbook(SPRE, data_only=True)
prowA = [list(r) for r in wsrcA["Participant Data"].iter_rows(values_only=True)][1:]
PDa = {}
for r in prowA:
    if r[0] is None:
        continue
    PDa[str(r[3]).strip()] = {"att": r[4], "acc": r[5], "score": r[6], "correct": r[7],
                              "incorrect": r[10], "unatt": r[12], "time": r[13]}
nm = [n for n in pre["P"]]
miss = [n for n in nm if n in PDa and int(PDa[n]["correct"]) not in nA]
check("PRE: jumlah BENAR tiap peserta", not miss, "gagal: %s" % (miss or "tidak ada"))
miss = [n for n in nm if n in PDa and float(PDa[n]["score"]) not in nA]
check("PRE: skor poin Wayground tiap peserta", not miss, "gagal: %s" % (miss or "tidak ada"))
miss = [n for n in nm if n in PDa and float(tsec(PDa[n]["time"])) not in nA]
check("PRE: waktu total tiap peserta (detik)", not miss, "gagal: %s" % (miss or "tidak ada"))
W("")

# ================= 6. MATRIKS WAKTU =================
W("-" * 100)
W("6. MATRIKS WAKTU PER BUTIR PER PESERTA")
W("-" * 100)
TT = {n: {int(k): v for k, v in d.items()} for n, d in TJ["TIME"].items()}
cells = [(n, q, TT[n][q]) for n in post["names"] for q in range(1, 21) if TT[n].get(q) is not None]
miss = [(n, q) for n, q, v in cells if float(v) not in nB]
check("POST: %d sel waktu per butir ada di workbook" % len(cells), not miss, "gagal: %d" % len(miss))
W("")

# ================= REKAP =================
W("=" * 100)
ok = sum(1 for n, o, d in res if o)
W("REKAP VERBATIM: %d dari %d pemeriksaan LULUS" % (ok, len(res)))
gg = [n for n, o, d in res if not o]
if gg:
    W("")
    W("BELUM TERPENUHI:")
    for n in gg:
        W("   - " + n)
W("=" * 100)

out = "\n".join(L)
io.open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "audit_verbatim.txt"),
        "w", encoding="utf-8").write(out)
print("selesai: %d/%d lulus" % (ok, len(res)))
