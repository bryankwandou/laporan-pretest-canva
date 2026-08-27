# -*- coding: utf-8 -*-
import openpyxl, json, re
from bs4 import BeautifulSoup
XL=r"E:/Download/pretestpelatihancanva25agustus2026-2026-08-25T09_22_13_634913-c1bee5.xlsx"
wb=openpyxl.load_workbook(XL,data_only=True)
ov=list(wb["Overview"].iter_rows(values_only=True))
hdr=ov[0]
names=[re.sub(r"\s*\(.*\)\s*$","",str(h)).strip() for h in hdr[12:49]]
Q=[]
for r in ov[1:21]:
    Q.append(dict(no=int(r[0]),text=str(r[1]).replace("\ufffd","->"),qtype=r[2],acc=int(str(r[4]).rstrip('%')),
        avgt=str(r[5]),correct=int(r[6]),incorrect=int(r[9]),unatt=int(r[11]),
        answers={names[i]:(str(r[12+i]).replace("\ufffd","->") if r[12+i] is not None else None) for i in range(37)}))
# participants
pd_=list(wb["Participant Data"].iter_rows(values_only=True))
P={}
for r in pd_[1:38]:
    n=str(r[3]).strip()
    h,m,s=[int(x) for x in str(r[13]).split(":")]
    P[n]=dict(name=n,att=int(r[4]),acc=int(str(r[5]).rstrip('%')),score=int(r[6]),correct=int(r[7]),
              incorrect=int(r[10]),unatt=int(r[12]),time_s=h*3600+m*60+s)
# time data
td=list(wb["Time Data"].iter_rows(values_only=True))
tnames=[str(x).strip() for x in td[0][5:42]]
TIME={n:{} for n in tnames}
for r in td[1:21]:
    qn=int(r[0])
    for i,n in enumerate(tnames):
        v=r[5+i]
        if v is None or str(v).strip() in ('None','-',''): TIME[n][qn]=None
        else:
            h,m,s=[int(x) for x in str(v).split(":")]; TIME[n][qn]=h*3600+m*60+s
# correctness from HTML
s=BeautifulSoup(open(r"E:/Download/Wayground 25 agustus 2026 canva wkri.html",encoding="utf-8",errors="ignore").read(),"html.parser")
basics=s.select('.player-basic-detail'); qrows=s.select('.player-question-detail')
CORR={}
for b,r in zip(basics,qrows):
    nm=b.select_one('.whitespace-nowrap').get_text(strip=True)
    cells=[]
    for c in r.select('[data-testid="time-taken-card"]'):
        cl=[x for x in c.get("class") if x.startswith("bg-")][0]
        cells.append({'bg-ds-success-500':'C','bg-red-light':'X','bg-ds-dark-100':'-'}[cl])
    CORR[nm]=cells
assert set(CORR)==set(P), (set(CORR)^set(P))
# derive key
for q in Q:
    votes={}
    for n,st in CORR.items():
        if st[q['no']-1]=='C':
            a=q['answers'].get(n); votes[a]=votes.get(a,0)+1
    q['key']=max(votes,key=votes.get) if votes else None
    q['key_votes']=votes
json.dump(dict(Q=Q,P=P,TIME=TIME,CORR=CORR,names=names),open("core.json","w",encoding="utf-8"),ensure_ascii=False,indent=1)
for q in Q: print(q['no'],q['acc'],'%',len(q['key_votes']),'|',(q['key'] or '')[:70])
